"""Tests de la plantilla y carga manual de UF (sprint 5).

Este módulo es el piloto del patrón que reusan los sprints 14 y 15, así que lo
que se fija acá vale para los tres: la plantilla dice qué falta, la carga es
idempotente por fecha, y los errores se informan por fila sin escribir nada a
medias.
"""
from datetime import date, timedelta
from decimal import Decimal as D
from io import BytesIO

import openpyxl
import pytest

from app.models.uf import UFDiaria
from app.services.importar_uf import (
    DIAS_PLANTILLA,
    UMBRAL_AVISO,
    cargar_desde_xlsx,
    estado_serie,
    generar_plantilla,
)

HOY = date(2026, 8, 21)


@pytest.fixture
def serie(db):
    """Una serie que llega hasta el 2026-09-09, como la real."""
    inicio = date(2026, 8, 1)
    db.add_all([
        UFDiaria(fecha=inicio + timedelta(days=i), valor=D("40800") + i)
        for i in range((date(2026, 9, 9) - inicio).days + 1)
    ])
    db.commit()
    return db


def _xlsx(filas, encabezados=("FECHA", "VALOR"), hoja="UF") -> bytes:
    libro = openpyxl.Workbook()
    libro.active.title = hoja
    libro.active.append(list(encabezados))
    for f in filas:
        libro.active.append(list(f))
    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------- estado


def test_serie_vacia_pide_cargarla(db):
    e = estado_serie(db, HOY)
    assert e.nivel == "vacia"
    assert e.ultima is None
    assert "vacía" in e.mensaje


def test_con_colchon_holgado_no_molesta(serie):
    e = estado_serie(serie, HOY)
    assert e.nivel == "ok"
    assert e.dias_de_colchon == 19
    assert e.ultima == date(2026, 9, 9)


def test_avisa_recien_a_tres_dias(serie):
    """El umbral es 3 y no 15 (D-008): con 15 quedaría prendido media vida."""
    assert estado_serie(serie, date(2026, 9, 5)).nivel == "ok", "4 días: todavía no"
    assert estado_serie(serie, date(2026, 9, 6)).nivel == "aviso", "3 días: ahora sí"
    assert estado_serie(serie, date(2026, 9, 9)).nivel == "aviso", "último día"


def test_la_serie_vencida_es_una_alerta_distinta(serie):
    e = estado_serie(serie, date(2026, 9, 10))
    assert e.nivel == "vencida"
    assert e.dias_de_colchon == -1
    assert "no se pueden valorizar" in e.mensaje.lower()


def test_el_umbral_es_el_documentado():
    assert UMBRAL_AVISO == 3


# --------------------------------------------------------------- plantilla


def test_la_plantilla_arranca_donde_termina_la_serie(serie):
    hoja = openpyxl.load_workbook(BytesIO(generar_plantilla(serie, HOY)))["UF"]

    assert [c.value for c in hoja[1]] == ["FECHA", "VALOR"]
    assert hoja.cell(2, 1).value == "2026-09-10", "el día siguiente al último cargado"
    assert hoja.cell(2, 2).value is None, "el valor viene en blanco"


def test_la_plantilla_trae_las_fechas_que_faltan(serie):
    hoja = openpyxl.load_workbook(BytesIO(generar_plantilla(serie, HOY)))["UF"]
    assert hoja.max_row == DIAS_PLANTILLA + 1


def test_con_la_serie_vacia_la_plantilla_arranca_hoy(db):
    """No tiene sentido pedir cuatro años de historia a mano."""
    hoja = openpyxl.load_workbook(BytesIO(generar_plantilla(db, HOY)))["UF"]
    assert hoja.cell(2, 1).value == HOY.isoformat()


def test_la_plantilla_trae_instrucciones(serie):
    libro = openpyxl.load_workbook(BytesIO(generar_plantilla(serie, HOY)))
    assert "Instrucciones" in libro.sheetnames
    texto = " ".join(str(c.value or "") for fila in libro["Instrucciones"] for c in fila)
    assert "VALOR" in texto and "no duplica" in texto


def test_la_plantilla_se_puede_cargar_de_vuelta(serie):
    """El ciclo completo: se baja, se rellena, se sube."""
    libro = openpyxl.load_workbook(BytesIO(generar_plantilla(serie, HOY)))
    hoja = libro["UF"]
    for fila in range(2, 5):
        hoja.cell(fila, 2, 41000 + fila)
    buffer = BytesIO()
    libro.save(buffer)

    r = cargar_desde_xlsx(serie, buffer.getvalue())

    assert (r.nuevas, r.actualizadas, r.errores) == (3, 0, [])


# --------------------------------------------------------------- carga


def test_carga_valores_nuevos(serie):
    r = cargar_desde_xlsx(serie, _xlsx([
        ("2026-09-10", 40900.5),
        ("2026-09-11", 40901.75),
    ]))

    assert (r.nuevas, r.actualizadas, r.sin_cambio) == (2, 0, 0)
    assert serie.get(UFDiaria, date(2026, 9, 10)).valor == D("40900.50")


def test_volver_a_subir_lo_mismo_no_duplica_ni_cambia(serie):
    """Subir meses solapados es lo normal, no un error."""
    archivo = _xlsx([("2026-09-01", 40831), ("2026-09-10", 40900)])

    primera = cargar_desde_xlsx(serie, archivo)
    segunda = cargar_desde_xlsx(serie, archivo)

    assert (primera.nuevas, primera.sin_cambio) == (1, 1)
    assert (segunda.nuevas, segunda.actualizadas, segunda.sin_cambio) == (0, 0, 2)


def test_un_valor_corregido_se_actualiza(serie):
    cargar_desde_xlsx(serie, _xlsx([("2026-09-10", 40900)]))
    r = cargar_desde_xlsx(serie, _xlsx([("2026-09-10", 40905)]))

    assert (r.nuevas, r.actualizadas) == (0, 1)
    assert serie.get(UFDiaria, date(2026, 9, 10)).valor == D("40905.00")


def test_las_filas_sin_valor_se_ignoran_sin_ruido(serie):
    """Son las de la plantilla que no se llenaron."""
    r = cargar_desde_xlsx(serie, _xlsx([
        ("2026-09-10", 40900),
        ("2026-09-11", None),
        ("2026-09-12", ""),
    ]))

    assert r.nuevas == 1
    assert r.errores == []


def test_acepta_las_dos_convenciones_de_numero(serie):
    r = cargar_desde_xlsx(serie, _xlsx([
        ("2026-09-10", "40.900,55"),
        ("2026-09-11", "40901.66"),
    ]))

    assert r.errores == []
    assert serie.get(UFDiaria, date(2026, 9, 10)).valor == D("40900.55")
    assert serie.get(UFDiaria, date(2026, 9, 11)).valor == D("40901.66")


def test_falta_una_columna_y_dice_cuales_encontro(serie):
    with pytest.raises(ValueError) as exc:
        cargar_desde_xlsx(serie, _xlsx([], encabezados=("FECHA", "MONTO")))

    mensaje = str(exc.value)
    assert "VALOR" in mensaje and "MONTO" in mensaje


def test_con_errores_no_se_carga_nada(serie):
    """Media serie subida sin saber cuál mitad es peor que no cargar."""
    antes = serie.query(UFDiaria).count()

    r = cargar_desde_xlsx(serie, _xlsx([
        ("2026-09-10", 40900),
        ("no es fecha", 40901),
        ("2026-09-12", "cero"),
    ]))

    assert len(r.errores) == 2
    assert r.nuevas == 0
    assert serie.query(UFDiaria).count() == antes
    assert "Fila 3" in r.errores[0] and "Fila 4" in r.errores[1]


def test_un_valor_negativo_es_un_error(serie):
    r = cargar_desde_xlsx(serie, _xlsx([("2026-09-10", -5)]))
    assert len(r.errores) == 1
    assert "positivo" in r.errores[0]


def test_la_misma_fecha_dos_veces_con_valores_distintos(serie):
    r = cargar_desde_xlsx(serie, _xlsx([
        ("2026-09-10", 40900),
        ("2026-09-10", 40999),
    ]))
    assert len(r.errores) == 1
    assert "dos veces" in r.errores[0]


def test_la_misma_fecha_dos_veces_con_el_mismo_valor_no_molesta(serie):
    r = cargar_desde_xlsx(serie, _xlsx([
        ("2026-09-10", 40900),
        ("2026-09-10", 40900),
    ]))
    assert r.errores == []
    assert r.nuevas == 1


# --------------------------------------------------------------- endpoints


def test_el_endpoint_de_estado(cliente, serie):
    datos = cliente.get("/api/uf/estado").json()
    assert datos["nivel"] in {"ok", "aviso", "vencida"}
    assert datos["ultima"] == "2026-09-09"


def test_la_plantilla_se_descarga_como_xlsx(cliente, serie):
    r = cliente.get("/api/uf/plantilla")

    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    assert openpyxl.load_workbook(BytesIO(r.content)).sheetnames == ["UF", "Instrucciones"]


def test_importar_por_el_endpoint(cliente, serie):
    archivo = _xlsx([("2026-09-10", 40900)])
    r = cliente.post(
        "/api/uf/importar",
        files={"archivo": ("uf.xlsx", archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert r.status_code == 200, r.text
    assert r.json()["nuevas"] == 1


def test_rechaza_un_archivo_que_no_es_xlsx(cliente, serie):
    r = cliente.post("/api/uf/importar", files={"archivo": ("datos.csv", b"a,b", "text/csv")})
    assert r.status_code == 400
    assert ".xlsx" in r.json()["detail"]
