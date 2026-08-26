"""La carga del historial de etapas de negocios.

Existe para desbloquear la proyeccion de plazos: hoy no hay ni un movimiento de
negocio registrado, asi que no se puede saber cuanto dura un negocio ni donde se
atasca. Esperar a que se acumule son meses; cargarlo hacia atras es una tarde.

Las cuatro reglas que se fijan son las que se acordaron antes de escribirla, y
cada una evita un dano concreto:

1. **No agenda proxima accion**, o cargar historia llenaria la bandeja de
   compromisos vencidos que nadie prometio.
2. **No hace retroceder la etapa vigente**, o cargar E1 y E2 de un negocio en E7
   lo bajaria a E2 y borraria el dato bueno con uno viejo.
3. **Recargar no duplica**, para poder iterar sin ensuciar la bitacora.
4. **No corrige una fecha de inicio si eso mueve plata**, y lo comprueba en vez de
   suponerlo.
"""
from datetime import date, datetime, timezone
from decimal import Decimal as D
from io import BytesIO

import openpyxl
import pytest
from sqlalchemy import select

from app.models.catalogo import EstadoNegocio, Etapa, ModeloNegocio
from app.models.movimiento import EntityType, Movimiento, TipoMovimiento
from app.models.negocio import Negocio, NegocioHito, Propiedad
from app.services.importar_historial import (
    ImportarHistorialError,
    importar_historial,
)
from app.services.plantilla_historial import (
    COLUMNAS_HISTORIAL,
    COLUMNAS_LIQUIDACIONES,
    HOJA_HISTORIAL,
    HOJA_LIQUIDACIONES,
    filas_de_liquidaciones,
    filas_del_historial,
    generar_plantilla,
)


@pytest.fixture(autouse=True)
def catalogo(db):
    db.add_all([
        Etapa(codigo="E1", nombre="Calificación", responsable="COMERCIAL", orden=1),
        Etapa(codigo="E2", nombre="Visita", responsable="COMERCIAL", orden=2),
        Etapa(codigo="E3", nombre="Promesa", responsable="OPERACIONES", orden=3),
        Etapa(codigo="E5", nombre="Escritura", responsable="OPERACIONES", orden=5),
    ])
    db.add_all([
        TipoMovimiento(codigo="NEG_E1", entity_type=EntityType.negocio, nombre="Paso a E1",
                       etapa_resultante="E1", orden=1, sla_es_habil=False, activo=True),
        TipoMovimiento(codigo="NEG_E2", entity_type=EntityType.negocio, nombre="Paso a E2",
                       etapa_resultante="E2", orden=2, sla_es_habil=False, activo=True),
        TipoMovimiento(codigo="NEG_E3", entity_type=EntityType.negocio, nombre="Paso a E3",
                       etapa_resultante="E3", orden=3, sla_es_habil=False, activo=True),
        TipoMovimiento(codigo="NEG_E5", entity_type=EntityType.negocio, nombre="Paso a E5",
                       etapa_resultante="E5", orden=5, sla_es_habil=False, activo=True),
    ])
    db.commit()
    return db


def _negocio(db, codigo, etapa="E2", hitos=None):
    prop = Propiedad(direccion=f"Calle {codigo}", comuna="Santiago")
    db.add(prop)
    n = Negocio(codigo=codigo, modelo=ModeloNegocio.MERCADO_PRIMARIO, propiedad=prop, etapa=etapa)
    n.hitos = hitos or [NegocioHito(fecha_inicio=date(2026, 3, 1), estado=EstadoNegocio.ACTIVO)]
    db.add(n)
    db.commit()
    return n


def _archivo(historial, liquidaciones=None) -> bytes:
    """Un .xlsx con las dos hojas, en el formato que espera el importador."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = HOJA_HISTORIAL
    for i, col in enumerate(COLUMNAS_HISTORIAL, start=1):
        hoja.cell(row=1, column=i, value=col.nombre)
    for f, fila in enumerate(historial, start=2):
        for i, valor in enumerate(fila, start=1):
            hoja.cell(row=f, column=i, value=valor)

    otra = libro.create_sheet(HOJA_LIQUIDACIONES)
    for i, col in enumerate(COLUMNAS_LIQUIDACIONES, start=1):
        otra.cell(row=1, column=i, value=col.nombre)
    for f, fila in enumerate(liquidaciones or [], start=2):
        for i, valor in enumerate(fila, start=1):
            otra.cell(row=f, column=i, value=valor)

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------- la plantilla


def test_la_plantilla_viene_pre_llenada_hasta_la_etapa_actual(db):
    """Un negocio en E3 trae E1, E2 y E3. Uno en E1 trae solo E1.

    Es el supuesto de secuencia: se asume que un negocio en E3 paso por las dos
    anteriores. Se declara en la guia del archivo, y las filas que no apliquen se
    borran.
    """
    _negocio(db, "A-1", etapa="E3")
    _negocio(db, "A-2", etapa="E1")

    filas = filas_del_historial(db)

    assert [(f[0], f[1]) for f in filas] == [
        ("A-1", "E1"), ("A-1", "E2"), ("A-1", "E3"),
        ("A-2", "E1"),
    ]
    # La fecha y la descripcion van vacias: son lo unico que hay que llenar.
    assert all(f[2] == "" and f[3] == "" for f in filas)


def test_las_fechas_de_referencia_van_solo_en_la_primera_fila(db):
    """Repetidas en cada fila se leerian como si fueran de esa etapa."""
    _negocio(db, "R-1", etapa="E3", hitos=[
        NegocioHito(fecha_inicio=date(2026, 1, 5), fecha_cierre=date(2026, 2, 20),
                    estado=EstadoNegocio.CERRADO),
    ])

    filas = filas_del_historial(db)

    assert filas[0][4] == "05-01-2026" and filas[0][5] == "20-02-2026"
    assert all(f[4] == "" and f[5] == "" for f in filas[1:])


def test_la_hoja_de_liquidaciones_trae_solo_las_de_fecha_unica(db):
    """Pedir la fecha de las que estan bien seria invitar a romperlas."""
    _negocio(db, "U-1", etapa="E5", hitos=[
        NegocioHito(nombre="PROMESA", fecha_inicio=date(2026, 2, 1),
                    fecha_cierre=date(2026, 2, 1), estado=EstadoNegocio.CERRADO),
        NegocioHito(nombre="ESCRITURA", fecha_inicio=date(2026, 3, 1),
                    fecha_cierre=date(2026, 4, 15), estado=EstadoNegocio.CERRADO),
    ])

    filas = filas_de_liquidaciones(db)

    assert [(f[0], f[1]) for f in filas] == [("U-1", "PROMESA")]


def test_la_plantilla_se_genera_y_la_lee_su_propio_importador(db):
    """El error de `D-046`: una plantilla que su propio importador rechaza.

    Ahi el encabezado quedo en dos filas copiado de otra plantilla y el importador
    leia la primera. Este test cierra ese circulo.
    """
    _negocio(db, "P-1", etapa="E2")

    contenido = generar_plantilla(db)
    libro = openpyxl.load_workbook(BytesIO(contenido))

    assert HOJA_HISTORIAL in libro.sheetnames
    assert HOJA_LIQUIDACIONES in libro.sheetnames
    # Sin fechas llenas no carga nada, pero tampoco falla: es el caso de bajarla y
    # subirla sin tocarla.
    resumen = importar_historial(db, contenido)
    assert resumen.total_movimientos == 0
    assert resumen.filas_sin_fecha == 2


# ------------------------------------------------------------------ la carga


def test_carga_los_movimientos_con_su_fecha_y_su_descripcion(db):
    n = _negocio(db, "C-1", etapa="E3")

    resumen = importar_historial(db, _archivo([
        ("C-1", "E1", "10-01-2026", "Vino por Instagram"),
        ("C-1", "E2", "22-01-2026", "Visitó el depto"),
        ("C-1", "E3", "05-02-2026", ""),
    ]))

    assert resumen.movimientos_creados == 3
    movimientos = db.scalars(
        select(Movimiento).where(Movimiento.entity_id == n.id).order_by(Movimiento.fecha)
    ).all()
    assert [m.etapa_resultante for m in movimientos] == ["E1", "E2", "E3"]
    assert [m.fecha.date() for m in movimientos] == [
        date(2026, 1, 10), date(2026, 1, 22), date(2026, 2, 5)
    ]
    assert movimientos[0].comentario == "Vino por Instagram"
    assert movimientos[2].comentario is None


def test_no_agenda_proxima_accion(db):
    """Cargar historia no puede llenar la bandeja de vencidos que nadie prometio."""
    _negocio(db, "S-1", etapa="E2")

    importar_historial(db, _archivo([("S-1", "E1", "10-01-2026", "")]))

    movimiento = db.scalar(select(Movimiento))
    assert movimiento.proximo_seguimiento is None


def test_no_hace_retroceder_la_etapa_del_negocio(db):
    """Cargar E1 de un negocio en E5 no puede bajarlo a E1.

    Registrar un movimiento mueve la etapa del negocio al del movimiento
    cronologicamente mas nuevo (`D-060`). En una carga historica eso borraria el
    dato bueno con uno viejo, asi que la etapa no se toca.
    """
    n = _negocio(db, "T-1", etapa="E5")

    importar_historial(db, _archivo([("T-1", "E1", "10-01-2026", "")]))

    db.refresh(n)
    assert n.etapa == "E5"


def test_recargar_no_duplica_y_corrige_la_fecha(db):
    n = _negocio(db, "D-1", etapa="E2")

    importar_historial(db, _archivo([("D-1", "E1", "10-01-2026", "primera version")]))
    resumen = importar_historial(db, _archivo([("D-1", "E1", "15-01-2026", "corregida")]))

    assert resumen.movimientos_creados == 0
    assert resumen.movimientos_actualizados == 1
    movimientos = db.scalars(
        select(Movimiento).where(Movimiento.entity_id == n.id)
    ).all()
    assert len(movimientos) == 1
    assert movimientos[0].fecha.date() == date(2026, 1, 15)
    assert movimientos[0].comentario == "corregida"


def test_las_filas_sin_fecha_se_ignoran_sin_fallar(db):
    """Es el caso normal: de un negocio se saben dos fechas y no las siete."""
    _negocio(db, "V-1", etapa="E3")

    resumen = importar_historial(db, _archivo([
        ("V-1", "E1", "10-01-2026", ""),
        ("V-1", "E2", "", ""),
        ("V-1", "E3", None, ""),
    ]))

    assert resumen.movimientos_creados == 1
    assert resumen.filas_sin_fecha == 2
    assert resumen.omitidas == []


def test_una_fila_mala_no_aborta_las_buenas(db):
    """Un archivo de 71 filas con tres errores tiene 68 filas buenas."""
    _negocio(db, "M-1", etapa="E2")

    resumen = importar_historial(db, _archivo([
        ("M-1", "E1", "10-01-2026", ""),
        ("NO-EXISTE", "E1", "10-01-2026", ""),
        ("M-1", "E9", "10-01-2026", ""),
        ("M-1", "E2", "no es una fecha", ""),
    ]))

    assert resumen.movimientos_creados == 1
    assert len(resumen.omitidas) == 3
    assert any("NO-EXISTE" in m for m in resumen.omitidas)
    assert any("E9" in m for m in resumen.omitidas)
    assert any("no se entiende la fecha" in m for m in resumen.omitidas)


def test_rechaza_una_fecha_futura(db):
    _negocio(db, "F-1", etapa="E2")
    manana = datetime.now(timezone.utc).date().replace(year=2099)

    resumen = importar_historial(db, _archivo([
        ("F-1", "E1", manana.strftime("%d-%m-%Y"), ""),
    ]))

    assert resumen.movimientos_creados == 0
    assert any("futuro" in m for m in resumen.omitidas)


def test_avisa_cuando_la_fecha_queda_antes_del_inicio_registrado(db):
    """No es un error: es el motivo por el que existe la hoja LIQUIDACIONES.

    En 7 liquidaciones la fecha de inicio esta mal --es la de cierre-- asi que las
    fechas reales van a caer antes. Se permiten y se listan.
    """
    _negocio(db, "AI-1", etapa="E2", hitos=[
        NegocioHito(fecha_inicio=date(2026, 6, 1), fecha_cierre=date(2026, 6, 1),
                    estado=EstadoNegocio.CERRADO),
    ])

    resumen = importar_historial(db, _archivo([("AI-1", "E1", "10-01-2026", "")]))

    assert resumen.movimientos_creados == 1, "se permite, no se rechaza"
    assert len(resumen.anteriores_al_inicio) == 1
    assert "AI-1" in resumen.anteriores_al_inicio[0]


# ------------------------------------------- la correccion de fechas de inicio


def test_corrige_la_fecha_de_inicio_cuando_no_mueve_plata(db):
    """Con `fecha_valorizacion` puesta, la UF ya esta congelada y no depende de
    `fecha_inicio`: corregirla es seguro."""
    n = _negocio(db, "L-1", etapa="E5", hitos=[
        NegocioHito(nombre=None, fecha_inicio=date(2026, 4, 12), fecha_cierre=date(2026, 4, 12),
                    fecha_valorizacion=date(2026, 4, 12), estado=EstadoNegocio.CERRADO),
    ])

    resumen = importar_historial(db, _archivo([], [("L-1", "ÚNICA", "01-02-2026")]))

    assert resumen.fechas_corregidas == 1
    db.refresh(n)
    assert n.hitos[0].fecha_inicio == date(2026, 2, 1)


def test_no_corrige_si_eso_moveria_la_plata(db):
    """Sin `fecha_valorizacion` ni valor manual, la UF sale de `fecha_inicio`.

    Cambiarla moveria el monto y la comision. La carga se niega y lo dice, en vez
    de alterar en silencio algo que ya funcionaba.
    """
    n = _negocio(db, "L-2", etapa="E5", hitos=[
        NegocioHito(nombre=None, fecha_inicio=date(2026, 4, 12), fecha_cierre=date(2026, 4, 12),
                    fecha_valorizacion=None, valor_clp_manual=None,
                    estado=EstadoNegocio.CERRADO),
    ])

    resumen = importar_historial(db, _archivo([], [("L-2", "ÚNICA", "01-02-2026")]))

    assert resumen.fechas_corregidas == 0
    assert len(resumen.no_corregidas_por_plata) == 1
    assert "L-2" in resumen.no_corregidas_por_plata[0]
    db.refresh(n)
    assert n.hitos[0].fecha_inicio == date(2026, 4, 12), "no se toco"


def test_el_valor_manual_tambien_protege_la_plata(db):
    """Con valor manual, la base no sale de la UF (`D-017`), asi que es seguro."""
    n = _negocio(db, "L-3", etapa="E5", hitos=[
        NegocioHito(nombre=None, fecha_inicio=date(2026, 4, 12), fecha_cierre=date(2026, 4, 12),
                    fecha_valorizacion=None, valor_clp_manual=D("100000000"),
                    estado=EstadoNegocio.CERRADO),
    ])

    resumen = importar_historial(db, _archivo([], [("L-3", "ÚNICA", "01-02-2026")]))

    assert resumen.fechas_corregidas == 1
    db.refresh(n)
    assert n.hitos[0].fecha_inicio == date(2026, 2, 1)


def test_no_acepta_un_inicio_posterior_al_cierre(db):
    _negocio(db, "L-4", etapa="E5", hitos=[
        NegocioHito(nombre=None, fecha_inicio=date(2026, 4, 12), fecha_cierre=date(2026, 4, 12),
                    fecha_valorizacion=date(2026, 4, 12), estado=EstadoNegocio.CERRADO),
    ])

    resumen = importar_historial(db, _archivo([], [("L-4", "ÚNICA", "01-06-2026")]))

    assert resumen.fechas_corregidas == 0
    assert any("posterior al cierre" in m for m in resumen.omitidas)


def test_un_archivo_sin_la_hoja_del_historial_se_rechaza(db):
    libro = openpyxl.Workbook()
    libro.active.title = "OTRA COSA"
    buffer = BytesIO()
    libro.save(buffer)

    with pytest.raises(ImportarHistorialError, match=HOJA_HISTORIAL):
        importar_historial(db, buffer.getvalue())


def test_un_encabezado_distinto_se_rechaza_con_el_detalle(db):
    """Bajar la plantilla de nuevo es la salida, y el mensaje lo dice."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = HOJA_HISTORIAL
    hoja.cell(row=1, column=1, value="Codigo")
    buffer = BytesIO()
    libro.save(buffer)

    with pytest.raises(ImportarHistorialError, match="Negocio"):
        importar_historial(db, buffer.getvalue())
