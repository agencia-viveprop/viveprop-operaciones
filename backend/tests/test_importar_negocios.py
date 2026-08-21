"""Tests de la carga masiva de negocios (sprints 14 y 15).

Las propiedades que protegen:

1. **Si hay un solo error no se escribe nada.** Media carga es peor que ninguna:
   nadie sabe cuál mitad quedó y hay que revisar fila por fila para averiguarlo.
2. **Cargar dos veces no duplica.** Es lo que permite corregir el archivo y
   volver a subirlo, que es como la gente realmente trabaja.
3. **El motor calcula las comisiones.** La plantilla no tiene columnas de
   comisión a propósito. Si alguna vez las tuviera, este test se caería.
"""
from decimal import Decimal as D
from io import BytesIO

import openpyxl
import pytest
from sqlalchemy import select

from app.models.negocio import Negocio, NegocioHito, Propiedad
from app.services.importar_negocios import (
    FILA_ENCABEZADO,
    ArchivoInvalido,
    cargar_desde_xlsx,
)
from app.services.plantilla_negocios import HOJA, NOMBRES, generar_plantilla

# Una fila mínima válida: solo las obligatorias.
BASE = {
    "CODIGO": "VVP-100",
    "DIRECCION": "Los Militares 5620",
    "COMUNA": "Las Condes",
    "MODELO": "MERCADO_PRIMARIO",
    "ESTADO": "ACTIVO",
    "FECHA_INICIO": "2026-03-01",
}


@pytest.fixture
def construir():
    """Arma un .xlsx con la forma de la plantilla: encabezado en la fila 2."""
    def _construir(filas: list[dict], columnas: list[str] | None = None) -> bytes:
        cols = columnas if columnas is not None else list(NOMBRES)
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = HOJA
        hoja.append(["grupo"] * len(cols))          # fila 1: los grupos
        hoja.append(cols)                            # fila 2: los nombres
        for f in filas:
            hoja.append([f.get(c) for c in cols])
        buffer = BytesIO()
        libro.save(buffer)
        return buffer.getvalue()
    return _construir


@pytest.fixture
def base(db, catalogos_sembrados, uf_cargada):
    """Catálogos, etapas y una serie de UF para que se pueda valorizar."""
    return db


# ------------------------------------------------------------- la plantilla


def test_la_plantilla_no_tiene_columnas_de_comision(base):
    """Si aparecieran, alguien escribiría un número y el motor dejaría de mandar."""
    libro = openpyxl.load_workbook(BytesIO(generar_plantilla(base)))
    encabezados = [c.value for c in libro[HOJA][FILA_ENCABEZADO] if c.value]

    prohibidas = [h for h in encabezados if "COMISION" in h or "REBATE_MONTO" in h]
    assert prohibidas == []


def test_la_plantilla_trae_los_codigos_de_la_base(base):
    """Se generan, no se escriben: una alianza nueva aparece sola."""
    libro = openpyxl.load_workbook(BytesIO(generar_plantilla(base)))
    hoja = libro["Valores válidos"]
    codigos = {hoja.cell(row=r, column=1).value for r in range(1, hoja.max_row + 1)}

    assert "ASSETPLAN" in codigos
    assert "INGEVEC" in codigos
    # La alianza inactiva no se ofrece: no tiene sentido cargar contra ella.
    assert "ANTIGUA" not in codigos


def test_la_plantilla_recien_bajada_se_puede_cargar(base):
    """No debe quejarse de columnas faltantes contra su propio formato."""
    resumen = cargar_desde_xlsx(base, generar_plantilla(base))

    # Vacía, así que no carga nada, pero el error es "no hay filas", no de formato.
    assert resumen.errores == ["El archivo no tiene ninguna fila con datos."]


def test_la_vuelta_completa_bajar_llenar_y_subir(base):
    """El test que importa: cualquier desajuste entre el generador y el parser
    aparece acá y en ningún otro lado.

    Los tests que arman el .xlsx a mano comparten la suposición de cómo es la
    plantilla; este usa la de verdad, con sus dos filas de encabezado, sus hojas
    extra y sus celdas combinadas.
    """
    libro = openpyxl.load_workbook(BytesIO(generar_plantilla(base)))
    hoja = libro[HOJA]
    encabezados = [c.value for c in hoja[FILA_ENCABEZADO]]

    fila = BASE | {
        "ALIANZA": "INGEVEC", "TIPO_OPERACION": "VENTA", "ETAPA": "E1",
        "VALOR_NEGOCIO": 1000, "MONEDA": "UF", "FECHA_VALORIZACION": "2026-01-02",
        "PCT_LADO_VENDEDOR": 2, "PCT_BROKER_VENDEDOR": 1, "PCT_VP_VENDEDOR": 1,
    }
    hoja.append([fila.get(h) for h in encabezados])

    buffer = BytesIO()
    libro.save(buffer)
    resumen = cargar_desde_xlsx(base, buffer.getvalue())

    assert resumen.errores == []
    assert (resumen.negocios_nuevos, resumen.hitos_nuevos) == (1, 1)
    negocio = base.scalar(select(Negocio).where(Negocio.codigo == "VVP-100"))
    assert negocio.etapa == "E1"
    assert negocio.alianza_id is not None
    assert negocio.hitos[0].comision_total == D("794712.60")


# ------------------------------------------------------------------- carga


def test_carga_una_fila_minima(base, construir):
    resumen = cargar_desde_xlsx(base, construir([BASE]))

    assert resumen.errores == []
    assert (resumen.negocios_nuevos, resumen.hitos_nuevos) == (1, 1)
    negocio = base.scalar(select(Negocio).where(Negocio.codigo == "VVP-100"))
    assert negocio.propiedad.comuna == "Las Condes"
    assert len(negocio.hitos) == 1


def test_el_codigo_repetido_son_hitos_del_mismo_negocio(base, construir):
    """Como VVP-3, que tiene PROMESA y ESCRITURA."""
    filas = [
        BASE | {"HITO": "PROMESA"},
        BASE | {"HITO": "ESCRITURA", "FECHA_INICIO": "2026-04-01"},
    ]
    resumen = cargar_desde_xlsx(base, construir(filas))

    assert resumen.errores == []
    assert (resumen.negocios_nuevos, resumen.hitos_nuevos) == (1, 2)
    negocio = base.scalar(select(Negocio).where(Negocio.codigo == "VVP-100"))
    assert {h.nombre for h in negocio.hitos} == {"PROMESA", "ESCRITURA"}


def test_las_tasas_se_escriben_en_porcentaje_y_se_guardan_como_fraccion(base, construir):
    """Pedir 0,0252 en una planilla es pedir que alguien se equivoque."""
    fila = BASE | {"PCT_LADO_VENDEDOR": 2, "PCT_EQUIPO": "2,52"}
    cargar_desde_xlsx(base, construir([fila]))

    hito = base.scalar(select(NegocioHito))
    assert hito.pct_lado_vendedor == D("0.02")
    assert hito.pct_equipo == D("0.0252")


def test_el_motor_calcula_las_comisiones(base, construir):
    """Nada de lo que el motor calcula viene del archivo."""
    fila = BASE | {
        "VALOR_NEGOCIO": 1000, "MONEDA": "UF", "FECHA_VALORIZACION": "2026-01-02",
        "PCT_LADO_VENDEDOR": 2, "PCT_BROKER_VENDEDOR": 1, "PCT_VP_VENDEDOR": 1,
    }
    resumen = cargar_desde_xlsx(base, construir([fila]))
    assert resumen.errores == []

    hito = base.scalar(select(NegocioHito))
    # UF del 2026-01-02 = 39.735,63 -> 1.000 UF = 39.735.630
    assert hito.uf_snapshot == D("39735.63")
    assert hito.valor_clp_calculado == D("39735630.00")
    assert hito.comision_total == D("794712.60")   # 2% de la base
    assert hito.comision_broker == D("397356.30")  # 1%
    assert hito.comision_vp_bruta == D("397356.30")


def test_la_propiedad_se_reutiliza(base, construir):
    """Dos negocios en la misma unidad no crean dos propiedades."""
    filas = [BASE, BASE | {"CODIGO": "VVP-101"}]
    cargar_desde_xlsx(base, construir(filas))

    assert len(base.execute(select(Propiedad)).scalars().all()) == 1


def test_las_fechas_aceptan_las_dos_convenciones(base, construir):
    filas = [
        BASE | {"FECHA_INICIO": "2026-03-01"},
        BASE | {"CODIGO": "VVP-101", "FECHA_INICIO": "01-03-2026"},
    ]
    resumen = cargar_desde_xlsx(base, construir(filas))

    assert resumen.errores == []
    fechas = {h.fecha_inicio.isoformat() for h in base.execute(select(NegocioHito)).scalars()}
    assert fechas == {"2026-03-01"}


# ------------------------------------------------------------ idempotencia


def test_cargar_dos_veces_actualiza_y_no_duplica(base, construir):
    archivo = construir([BASE])
    cargar_desde_xlsx(base, archivo)
    segunda = cargar_desde_xlsx(base, archivo)

    assert (segunda.negocios_nuevos, segunda.hitos_nuevos) == (0, 0)
    assert (segunda.negocios_actualizados, segunda.hitos_actualizados) == (1, 1)
    assert len(base.execute(select(Negocio)).scalars().all()) == 1
    assert len(base.execute(select(NegocioHito)).scalars().all()) == 1


def test_una_correccion_se_aplica_al_volver_a_subir(base, construir):
    cargar_desde_xlsx(base, construir([BASE | {"COMUNA": "Las Condes"}]))
    cargar_desde_xlsx(base, construir([BASE | {"CORREDOR_AGENTE": "Nuevo corredor"}]))

    negocio = base.scalar(select(Negocio).where(Negocio.codigo == "VVP-100"))
    assert negocio.corredor_agente == "Nuevo corredor"


def test_no_borra_el_hito_que_el_archivo_no_menciona(base, construir):
    """Un archivo incompleto no puede ser pérdida de datos."""
    cargar_desde_xlsx(base, construir([
        BASE | {"HITO": "PROMESA"},
        BASE | {"HITO": "ESCRITURA"},
    ]))
    cargar_desde_xlsx(base, construir([BASE | {"HITO": "PROMESA"}]))

    negocio = base.scalar(select(Negocio).where(Negocio.codigo == "VVP-100"))
    assert {h.nombre for h in negocio.hitos} == {"PROMESA", "ESCRITURA"}


# ------------------------------------------------------- nada a medias


def test_un_error_en_una_fila_impide_toda_la_carga(base, construir):
    """La fila 1 es perfecta y tampoco se carga: es el punto."""
    filas = [BASE, BASE | {"CODIGO": "VVP-101", "MODELO": "INVENTADO"}]
    resumen = cargar_desde_xlsx(base, construir(filas))

    assert len(resumen.errores) == 1
    assert "INVENTADO" in resumen.errores[0]
    assert base.execute(select(Negocio)).first() is None


def test_informa_todos_los_problemas_juntos(base, construir):
    """Quien corrige el archivo prefiere verlos todos de una vez."""
    fila = {"CODIGO": "VVP-100", "MODELO": "MERCADO_PRIMARIO", "ESTADO": "CERRADO"}
    resumen = cargar_desde_xlsx(base, construir([fila]))

    juntos = " | ".join(resumen.errores)
    assert "DIRECCION" in juntos
    assert "COMUNA" in juntos
    assert "FECHA_INICIO" in juntos
    assert "FECHA_CIERRE" in juntos  # CERRADO sin fecha de cierre


@pytest.mark.parametrize("cambio, trozo", [
    ({"ESTADO": "CERRADO"}, "FECHA_CIERRE"),
    ({"FECHA_CIERRE": "2026-01-01"}, "anterior a FECHA_INICIO"),
    ({"VALOR_NEGOCIO": 100}, "MONEDA"),
    ({"VALOR_NEGOCIO": 100, "MONEDA": "UF"}, "FECHA_VALORIZACION"),
    ({"ETAPA": "E99"}, "no existe"),
    ({"ALIANZA": "NO_EXISTE"}, "Valores válidos"),
    ({"FECHA_INICIO": "no es fecha"}, "fecha inválida"),
    ({"VALOR_NEGOCIO": "abc", "MONEDA": "CLP"}, "número inválido"),
])
def test_las_incoherencias_se_detectan(base, construir, cambio, trozo):
    resumen = cargar_desde_xlsx(base, construir([BASE | cambio]))

    assert any(trozo in e for e in resumen.errores), resumen.errores
    assert base.execute(select(Negocio)).first() is None


def test_el_mismo_hito_dos_veces_en_el_archivo_es_un_error(base, construir):
    resumen = cargar_desde_xlsx(base, construir([BASE, BASE]))

    assert any("aparece dos veces" in e for e in resumen.errores)
    assert base.execute(select(Negocio)).first() is None


def test_dos_filas_del_mismo_negocio_no_pueden_discrepar(base, construir):
    """No hay forma de saber cuál dirección gana."""
    filas = [
        BASE | {"HITO": "PROMESA"},
        BASE | {"HITO": "ESCRITURA", "DIRECCION": "Otra calle 123"},
    ]
    resumen = cargar_desde_xlsx(base, construir(filas))

    assert any("DIRECCION" in e and "iguales" in e for e in resumen.errores)
    assert base.execute(select(Negocio)).first() is None


def test_un_archivo_sin_las_columnas_de_la_plantilla_se_rechaza(base, construir):
    with pytest.raises(ArchivoInvalido, match="CODIGO"):
        cargar_desde_xlsx(base, construir([{}], columnas=["OTRA_COSA", "MAS"]))


def test_un_archivo_que_no_es_excel_se_rechaza(base):
    with pytest.raises(ArchivoInvalido, match="No se pudo leer"):
        cargar_desde_xlsx(base, b"esto no es un xlsx")


def test_las_filas_vacias_de_la_plantilla_se_ignoran(base, construir):
    resumen = cargar_desde_xlsx(base, construir([BASE, {}, {}]))

    assert resumen.errores == []
    assert resumen.hitos_nuevos == 1


# --------------------------------------------------------------- endpoints


def test_el_endpoint_entrega_la_plantilla(cliente, catalogos_sembrados):
    r = cliente.get("/api/negocios/plantilla")

    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "plantilla-negocios.xlsx" in r.headers["content-disposition"]
    libro = openpyxl.load_workbook(BytesIO(r.content))
    assert HOJA in libro.sheetnames


def test_el_endpoint_carga(cliente, catalogos_sembrados, uf_cargada, construir):
    archivo = construir([BASE])
    r = cliente.post(
        "/api/negocios/importar",
        files={"archivo": ("carga.xlsx", archivo,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert r.status_code == 200
    assert r.json()["negocios_nuevos"] == 1


def test_el_endpoint_devuelve_los_errores_con_200(cliente, catalogos_sembrados, construir):
    """Son decenas de mensajes por fila y el front los lista: no caben en un 400."""
    archivo = construir([{"CODIGO": "VVP-1"}])
    r = cliente.post(
        "/api/negocios/importar",
        files={"archivo": ("carga.xlsx", archivo, "application/octet-stream")},
    )

    assert r.status_code == 200
    assert len(r.json()["errores"]) > 0
    assert r.json()["negocios_nuevos"] == 0


def test_el_endpoint_rechaza_un_archivo_con_otro_formato(cliente, catalogos_sembrados):
    """Esto sí es 400: no hay nada que listar, el archivo está mal."""
    r = cliente.post(
        "/api/negocios/importar",
        files={"archivo": ("cualquiera.txt", b"no soy un xlsx", "text/plain")},
    )

    assert r.status_code == 400
    assert "No se pudo leer" in r.json()["detail"]
