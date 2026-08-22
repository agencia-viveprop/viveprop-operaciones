"""Que la estructura que se muestra en pantalla sea la que la carga exige de verdad.

Las dos cargas masivas pedían un `.xlsx` sin decir en ninguna parte qué columnas
esperaban. La de negocios tenía plantilla, así que la respuesta estaba dentro de un
archivo que había que bajar y abrir; la de canjes no tenía ni eso. Ahora las dos
exponen su estructura y la pantalla la muestra.

**Lo que estos tests cuidan es que no diverja.** Una pantalla que describe una
columna que la carga ya no pide --o que se calla una nueva-- es peor que no
describir nada, porque se le cree. Los dos casos se comprueban contra la fuente
real: `plantilla_negocios.COLUMNAS`, que es lo que pinta el Excel, e
`importar_canjes.COLUMNAS_REQUERIDAS`, que es lo que la carga verifica.
"""
from io import BytesIO

import openpyxl
import pytest

from app.services.importar_canjes import COLUMNAS_REQUERIDAS
from app.services.plantilla_canjes import (
    COLUMNAS as COLUMNAS_CANJES,
    estructura_importacion,
    generar_plantilla as plantilla_canjes,
)
from app.services.plantilla_negocios import (
    COLUMNAS as COLUMNAS_NEGOCIOS,
    NOMBRES,
    estructura_plantilla,
)


# ------------------------------------------------------------------- canjes


def test_las_columnas_descritas_son_las_que_la_carga_exige():
    """El test central de este archivo.

    `importar_canjes` rechaza el archivo si le falta alguna de
    `COLUMNAS_REQUERIDAS`. La descripción vive en otro módulo, así que sin esto
    alguien puede agregar una columna a la carga y dejar la pantalla explicando
    las de antes, sin que nada falle.
    """
    descritas = [c.nombre for c in COLUMNAS_CANJES]

    # En el mismo orden, no solo las mismas: la plantilla se baja para comparar
    # encabezados contra un archivo que falla, y en distinto orden no sirve.
    assert descritas == list(COLUMNAS_REQUERIDAS), (
        "plantilla_canjes.COLUMNAS y importar_canjes.COLUMNAS_REQUERIDAS "
        f"divergieron en: {set(descritas) ^ set(COLUMNAS_REQUERIDAS)}"
    )


def test_la_estructura_de_canjes_agrupa_las_16_columnas():
    e = estructura_importacion()

    assert e.total_columnas == len(COLUMNAS_REQUERIDAS) == 16
    # Los grupos son solo para la pantalla y tienen que ser contiguos, así que
    # salen del orden del export: LINK_PROPIEDAD viene al final, después del
    # valor, y por eso queda solo en el suyo.
    assert [g.nombre for g in e.grupos] == [
        "Identificación", "Corredores", "Propiedad", "Valor", "Enlace",
    ]
    # Todas obligatorias: el archivo sale de una query, no se llena a mano.
    assert all(c.obligatoria for g in e.grupos for c in g.columnas)
    # Y todas con algo que decir: una fila sin ayuda no explica nada.
    assert all(c.ayuda.strip() for g in e.grupos for c in g.columnas)


def test_los_valores_de_canjes_salen_de_los_mapas_de_la_carga():
    """No escritos a mano: son los mismos que `_mapear` acepta (`D-048`)."""
    e = estructura_importacion()
    por_columna = {v.columna: v.valores for v in e.valores}

    assert por_columna["ESTADO"] == ["Activo", "Cancelado"]
    assert "En revisión" in por_columna["ETAPA"], "van con acento, como los escribe Dataprop"
    assert por_columna["TIPO_OPERACION"] == ["Venta", "Arriendo", "Otro/Desconocido"]
    assert por_columna["MONEDA_VALOR"] == ["CLP", "UF", "Otra"]


def test_la_plantilla_de_canjes_trae_los_encabezados_exactos():
    """En la fila 1, que es donde `importar_canjes` los busca.

    La plantilla de negocios tiene dos filas de encabezado --su carga lee la
    segunda-- y copiar ese estilo acá producía una plantilla que el propio
    cargador rechazaba por columnas faltantes.
    """
    libro = openpyxl.load_workbook(BytesIO(plantilla_canjes()))
    hoja = libro["CANJES"]

    assert [c.value for c in hoja[1] if c.value] == list(COLUMNAS_REQUERIDAS)
    # Ninguna fila de datos: es una plantilla, no un export.
    assert hoja.max_row == 1
    assert "Valores válidos" in libro.sheetnames


def test_la_plantilla_de_canjes_la_acepta_su_propia_carga(db):
    """La plantilla vacía tiene que pasar la validación de columnas.

    Si no, no serviría para lo único que se baja: comparar encabezados cuando la
    carga se queja. Que no cargue ninguna fila es lo esperado --está vacía--; lo
    que no puede pasar es que se rechace por columnas faltantes.
    """
    from app.services.importar_canjes import importar_canjes

    resumen = importar_canjes(db, plantilla_canjes())

    assert resumen.errores == []
    assert (resumen.nuevas, resumen.actualizadas, resumen.ignoradas) == (0, 0, 0)


# ----------------------------------------------------------------- negocios


def test_la_estructura_de_negocios_sale_de_las_columnas_de_la_plantilla(db, catalogos_sembrados):
    e = estructura_plantilla(db)

    assert e.total_columnas == len(COLUMNAS_NEGOCIOS) == 32
    # El mismo orden que el Excel, no uno alfabético: quien llena la fila la lee
    # de izquierda a derecha.
    aplanadas = [c.nombre for g in e.grupos for c in g.columnas]
    assert aplanadas == list(NOMBRES)
    assert all(c.ayuda.strip() for g in e.grupos for c in g.columnas)


def test_marca_como_obligatorias_las_mismas_que_la_plantilla(db, catalogos_sembrados):
    e = estructura_plantilla(db)
    obligatorias = {c.nombre for g in e.grupos for c in g.columnas if c.obligatoria}

    assert obligatorias == {c.nombre for c in COLUMNAS_NEGOCIOS if c.obligatoria}
    # Las que sin ellas no se puede armar ni el negocio ni el hito.
    assert obligatorias == {
        "CODIGO", "DIRECCION", "COMUNA", "MODELO", "ESTADO", "FECHA_INICIO",
    }


def test_los_valores_de_negocios_se_leen_de_la_base(db, catalogos_sembrados):
    """Una alianza nueva aparece sola, igual que en la hoja de la plantilla."""
    from app.models.catalogo import Catalogo

    e = estructura_plantilla(db)
    antes = {v.columna: list(v.valores) for v in e.valores}
    assert "ASSETPLAN" in antes["ALIANZA"]
    assert "NUEVA_ALIANZA" not in antes["ALIANZA"]

    db.add(Catalogo(tipo="alianza", codigo="NUEVA_ALIANZA", nombre="Nueva", orden=99))
    db.commit()

    despues = {v.columna: list(v.valores) for v in estructura_plantilla(db).valores}
    assert "NUEVA_ALIANZA" in despues["ALIANZA"]


# ----------------------------------------------------------------- endpoints


@pytest.mark.parametrize(
    "url,columnas",
    [
        ("/api/canjes/plantilla/estructura", 16),
        ("/api/negocios/plantilla/estructura", 32),
    ],
)
def test_los_endpoints_devuelven_la_estructura(cliente, catalogos_sembrados, url, columnas):
    r = cliente.get(url)

    assert r.status_code == 200, r.text
    cuerpo = r.json()
    assert sum(len(g["columnas"]) for g in cuerpo["grupos"]) == columnas
    # Lo que la pantalla necesita para explicar el archivo, no solo la lista.
    assert cuerpo["origen"] and cuerpo["fila"] and cuerpo["notas"]


def test_el_endpoint_baja_la_plantilla_de_canjes(cliente):
    r = cliente.get("/api/canjes/plantilla")

    assert r.status_code == 200
    assert "plantilla-canjes.xlsx" in r.headers["content-disposition"]
    libro = openpyxl.load_workbook(BytesIO(r.content))
    assert [c.value for c in libro["CANJES"][1] if c.value] == list(COLUMNAS_REQUERIDAS)


def test_plantilla_no_se_confunde_con_un_id_de_canje(cliente):
    """La ruta va antes de "/{canje_id}", que si no se tragaría «plantilla».

    Sin el orden correcto esto daría un 422 intentando parsear «plantilla» como
    entero, que es el error que ya se había cometido con «/bandeja».
    """
    assert cliente.get("/api/canjes/plantilla").status_code == 200
    assert cliente.get("/api/canjes/plantilla/estructura").status_code == 200
