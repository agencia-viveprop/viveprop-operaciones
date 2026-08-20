"""Carga la serie de UF desde la hoja `UF` del Excel de operaciones.

Uso:
    python -m app.scripts.cargar_uf <ruta_al_xlsx> [desde_aaaa-mm-dd]

Por defecto carga desde 2022-11-01: el canje mas antiguo es del 2022-11-29 y
todo lo anterior (la serie arranca en 1977) es peso muerto.

Es un upsert por fecha, asi que correrlo dos veces no duplica.
"""
import sys
from datetime import date, datetime

import openpyxl
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.db import SessionLocal, engine
from app.models.uf import UFDiaria

DESDE_POR_DEFECTO = date(2022, 11, 1)
COL_FECHA = 3
COL_VALOR = 4


def leer_hoja(ruta: str, desde: date) -> list[dict]:
    hoja = openpyxl.load_workbook(ruta, data_only=True)["UF"]
    filas = []
    for n in range(1, hoja.max_row + 1):
        f = hoja.cell(n, COL_FECHA).value
        v = hoja.cell(n, COL_VALOR).value
        if not isinstance(f, datetime) or v is None:
            continue
        if f.date() < desde:
            continue
        filas.append({"fecha": f.date(), "valor": round(float(v), 2)})
    return filas


def cargar(filas: list[dict]) -> int:
    if not filas:
        return 0
    with SessionLocal() as db:
        stmt = pg_insert(UFDiaria).values(filas)
        stmt = stmt.on_conflict_do_update(
            index_elements=["fecha"],
            set_={"valor": stmt.excluded.valor, "actualizado_en": stmt.excluded.actualizado_en},
        )
        db.execute(stmt)
        db.commit()
    return len(filas)


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(1)
    ruta = sys.argv[1]
    desde = date.fromisoformat(sys.argv[2]) if len(sys.argv) > 2 else DESDE_POR_DEFECTO

    filas = leer_hoja(ruta, desde)
    print(f"Leidas {len(filas)} filas desde {desde.isoformat()} (host {engine.url.host})")
    if filas:
        print(f"  rango: {filas[0]['fecha']} -> {filas[-1]['fecha']}")
    n = cargar(filas)
    print(f"Cargadas/actualizadas: {n}")


if __name__ == "__main__":
    main()
