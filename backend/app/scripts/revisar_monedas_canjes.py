"""Genera el Excel de revisión de monedas de `valor_prop` en canjes.

**Por qué existe.** La moneda de «Valor propiedad» está mal etiquetada en 139 de
los 297 canjes migrados: hay arriendos en Vitacura guardados como "70 CLP" y
ventas en Providencia como "320.000.000 UF". Es el motivo por el que ese campo se
descartó como fuente de plata (`D-054`), y sin él no se puede calcular ninguna
comisión de canjes.

**La magnitud dice la verdad y la etiqueta no.** Las dos escalas están separadas
por cuatro órdenes de magnitud --una venta en UF anda en miles, la misma en pesos
en cientos de millones-- así que clasificar por monto no es adivinar. Medido sobre
los 297: la regla resuelve 288 sin ambigüedad y deja 9 que necesitan una persona,
porque su monto no funciona en ninguna de las dos monedas.

**No escribe nada en la base.** Genera un archivo para revisar. Aplicar los
cambios es un paso aparte y deliberado: corregir 139 filas a partir de una
inferencia, sin que nadie las haya mirado, es exactamente lo que no se quiere.

Se corre desde `backend/` con `python -m app.scripts.revisar_monedas_canjes`.
"""
import os
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

RAIZ = Path(__file__).resolve().parent.parent.parent
load_dotenv(RAIZ / ".env")

# Los cortes de la regla. Entre ellos no se afirma nada.
VENTA_UF_HASTA = 1_000_000
VENTA_CLP_DESDE = 20_000_000
ARRIENDO_UF_HASTA = 1_000
ARRIENDO_CLP_DESDE = 100_000
ARRIENDO_CLP_HASTA = 20_000_000


def inferir(op: str, valor: float) -> str | None:
    """La moneda real según la magnitud. `None` = no se puede afirmar."""
    if valor is None or valor <= 0:
        return None
    if op == "VENTA":
        if valor < VENTA_UF_HASTA:
            return "UF"
        if valor >= VENTA_CLP_DESDE:
            return "CLP"
        return None
    if op == "ARRIENDO":
        if valor < ARRIENDO_UF_HASTA:
            return "UF"
        if ARRIENDO_CLP_DESDE <= valor <= ARRIENDO_CLP_HASTA:
            return "CLP"
        return None
    return None


e = create_engine(os.environ["DATABASE_URL"])
with e.begin() as c:
    uf_hoy = c.execute(
        text("select valor from uf_diaria order by fecha desc limit 1")
    ).scalar()
    fecha_uf = c.execute(
        text("select fecha from uf_diaria order by fecha desc limit 1")
    ).scalar()
    filas = c.execute(text("""
        select id, estado::text, etapa::text, tipo_operacion::text, tipo_inmueble,
               comuna, direccion, valor_prop, moneda_valor::text
        from canjes order by id
    """)).all()

ambiguos, a_corregir, correctos = [], [], []
for f in filas:
    (id_, estado, etapa, op, inmueble, comuna, direccion, valor, moneda) = f
    real = inferir(op, float(valor) if valor is not None else None)
    fila = {
        "id": id_, "estado": estado, "etapa": etapa, "op": op,
        "inmueble": inmueble, "comuna": comuna, "direccion": direccion,
        "valor": valor, "moneda": moneda, "propuesta": real,
    }
    if real is None:
        ambiguos.append(fila)
    elif real != moneda:
        a_corregir.append(fila)
    else:
        correctos.append(fila)

print(f"correctos={len(correctos)}  a_corregir={len(a_corregir)}  ambiguos={len(ambiguos)}")

# --------------------------------------------------------------- el archivo

INDIGO = PatternFill("solid", fgColor="3D3EA8")
AMARILLO = PatternFill("solid", fgColor="FEF0F0")
GRIS_FONDO = PatternFill("solid", fgColor="EDEDF9")
BLANCO = Font(color="FFFFFF", bold=True)
GRIS = Font(color="6B7280", italic=True)

COLUMNAS = [
    ("Canje", 9, False),
    ("Estado", 12, False),
    ("Operación", 11, False),
    ("Tipo inmueble", 17, False),
    ("Comuna", 16, False),
    ("Valor guardado", 17, False),
    ("Moneda actual", 14, False),
    ("Moneda correcta", 16, True),
    ("Equivale en CLP", 18, False),
    ("¿Se ve razonable?", 19, False),
]

libro = openpyxl.Workbook()
hoja = libro.active
hoja.title = "REVISAR"


def escribir_encabezado(h, columnas):
    for i, (nombre, ancho, editable) in enumerate(columnas, start=1):
        celda = h.cell(row=1, column=i, value=nombre)
        celda.fill = INDIGO if editable else GRIS_FONDO
        celda.font = BLANCO if editable else GRIS
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.column_dimensions[get_column_letter(i)].width = ancho
    h.freeze_panes = h.cell(row=2, column=1)


escribir_encabezado(hoja, COLUMNAS)

# Los ambiguos primero: son los que necesitan una decisión de verdad.
fila_n = 2
for grupo, marca in ((ambiguos, True), (a_corregir, False)):
    for f in grupo:
        valor = float(f["valor"]) if f["valor"] is not None else None
        propuesta = f["propuesta"]
        # El equivalente en pesos solo se puede escribir si hay propuesta.
        if propuesta == "CLP":
            equivale = valor
        elif propuesta == "UF" and valor is not None:
            equivale = round(valor * float(uf_hoy))
        else:
            equivale = None

        datos = [
            f["id"], f["estado"], f["op"], f["inmueble"], f["comuna"],
            valor, f["moneda"], propuesta or "", equivale,
            "" if propuesta else "REVISAR: el monto no funciona en ninguna moneda",
        ]
        for i, valor_celda in enumerate(datos, start=1):
            celda = hoja.cell(row=fila_n, column=i, value=valor_celda)
            if marca:
                celda.fill = AMARILLO
            if i in (6, 9):
                celda.number_format = "#,##0"
        fila_n += 1

# --------------------------------------------------------- la hoja de guia
guia = libro.create_sheet("Instrucciones")
guia.column_dimensions["A"].width = 26
guia.column_dimensions["C"].width = 96

texto = [
    ("Qué es esto", "La lista de canjes cuya moneda de «Valor propiedad» quedó mal etiquetada en el origen. Una propiedad no vale 70 pesos ni 320 millones de UF, así que la magnitud dice la verdad y la etiqueta no."),
    ("Qué hay que hacer", "Revisar la columna «Moneda correcta». Viene con la propuesta ya puesta: si estás de acuerdo, no toques nada. Si no, escribí CLP o UF."),
    ("Las filas amarillas", f"Son {len(ambiguos)} y vienen SIN propuesta: su monto no funciona en ninguna de las dos monedas, así que probablemente le falten o le sobren ceros. Esas necesitan tu decisión, y puede que además haya que corregir el monto."),
    ("Cuántas son", f"{len(a_corregir)} para corregir + {len(ambiguos)} ambiguas. Las otras {len(correctos)} ya están bien y no están en esta lista."),
    ("La regla que se aplicó", f"Venta: bajo {VENTA_UF_HASTA:,} es UF, sobre {VENTA_CLP_DESDE:,} es CLP. Arriendo: bajo {ARRIENDO_UF_HASTA:,} es UF, entre {ARRIENDO_CLP_DESDE:,} y {ARRIENDO_CLP_HASTA:,} es CLP. Entre esos rangos no se afirma nada."),
    ("El equivalente en CLP", f"Calculado con la UF de {fecha_uf:%d-%m-%Y} (${float(uf_hoy):,.2f}). Es para que puedas juzgar si el monto es plausible; no se guarda en ninguna parte."),
    ("Ojo con el alcance", "Este archivo se generó desde la base de desarrollo, que puede estar atrasada respecto de producción. Si en producción hay canjes más nuevos, no están acá."),
    ("Qué NO hace este archivo", "Nada por sí solo. Es para revisar. Aplicar los cambios es un paso aparte."),
]
f = 1
for titulo, cuerpo in texto:
    guia.cell(row=f, column=1, value=titulo).font = Font(bold=True)
    guia.cell(row=f, column=3, value=cuerpo).alignment = Alignment(wrap_text=True, vertical="top")
    f += 2

destino = RAIZ.parent / "Archivos" / "revision-monedas-canjes.xlsx"
libro.save(destino)
print("guardado en", destino)
