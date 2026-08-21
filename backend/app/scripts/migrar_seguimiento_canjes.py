"""Migra el seguimiento operativo de canjes desde el Excel a `movimientos`.

Uso:
    python -m app.scripts.migrar_seguimiento_canjes <ruta_al_xlsx> [--reemplazar]

**El problema de las fechas, y qué se hizo.** La hoja registra *qué* pasos se
completaron —287 marcas de "✓ Sí" repartidas en diez columnas— pero **no cuándo**
se completó cada uno. Las fechas que hay son de otra cosa: `Fecha último update`
(69 filas), `Última gestión solicitante` (91) y `Última gestión propietario` (64).

Se decidió migrar la estructura y no inventar fechas por paso. Cada movimiento
de un canje recibe la **mejor fecha real disponible para ese canje**, según el
lado al que pertenece el paso, y su comentario dice explícitamente que viene de
la migración. Ninguna fecha es inventada; lo que es aproximado es la
correspondencia entre la fecha y el paso, y queda dicho en cada fila.

La alternativa era un solo movimiento por canje resumiendo todo. Se descartó
porque perdería *cuáles* pasos están hechos, que es justamente lo que hace falta
para seguir desde donde se quedó.
"""
import sys
from collections import Counter
from datetime import datetime, time, timezone

import openpyxl
from openpyxl.utils import column_index_from_string as ci
from sqlalchemy import delete, select

from app.db import SessionLocal, engine
from app.models.canje import Canje
from app.models.movimiento import EntityType, Movimiento, TipoMovimiento
from app.models.usuario import Usuario

HOJA = "✅ Seguimiento Operativo"
FILA_PRIMERA = 4

COL_ID = "B"
COL_FECHA_UPDATE = "H"
COL_OPERADOR = "J"
COL_OBSERVACIONES = "L"
COL_GESTION_SOLICITANTE = "R"
COL_GESTION_PROPIETARIO = "V"
COL_ACUERDO_SOLICITANTE = "AA"
COL_ACUERDO_PROPIETARIO = "AB"
COL_TIPO_CANCELACION = "AG"
COL_MOTIVO_CANCELACION = "AH"

# columna -> (codigo de tipo, de que lado es el paso)
# El lado decide qué fecha se usa: la gestión del solicitante o la del
# propietario, que son las dos únicas fechas de contacto que la hoja registra.
CHECKLIST = {
    "M": ("CLIENTE_CALIFICADO", "solicitante"),
    "N": ("PROPIEDAD_DISPONIBLE", "propietario"),
    "O": ("WA_CONFIRM_SOLICITANTE", "solicitante"),
    "P": ("EMAIL_REGISTRO_SOLICITANTE", "solicitante"),
    "S": ("WA_CONFIRM_PROPIETARIO", "propietario"),
    "T": ("EMAIL_REGISTRO_PROPIETARIO", "propietario"),
    "W": ("MANDATO_FIRMADO", "propietario"),
    "X": ("VALIDACION_SOLICITANTE", "solicitante"),
    "Y": ("VALIDACION_PROPIETARIO", "propietario"),
    "Z": ("ACUERDO_FIRMADO", "acuerdo"),
}

NOTA = "Migrado del Excel — fecha aproximada"
SIN_TIPO = {"N/A", "NA", "-", ""}


def _texto(v):
    if v is None:
        return None
    t = str(v).replace("\n", " ").strip()
    return t or None


def _fecha(v) -> datetime | None:
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    if hasattr(v, "year"):
        return datetime.combine(v, time(12, 0), tzinfo=timezone.utc)
    return None


def _es_si(v) -> bool:
    return "Sí" in str(v) or "Si" in str(v)


def _es_no(v) -> bool:
    return "✗" in str(v) or str(v).strip().lower().startswith("no")


class Migracion:
    def __init__(self, db, hoja, autor_id: int):
        self.db = db
        self.hoja = hoja
        self.autor_id = autor_id
        self.tipos = {
            t.codigo: t
            for t in db.scalars(
                select(TipoMovimiento).where(TipoMovimiento.entity_type == EntityType.canje)
            ).all()
        }
        self.canjes = set(db.scalars(select(Canje.id)).all())
        self.creados = Counter()
        self.sin_canje: list[int] = []
        self.sin_fecha: list[int] = []
        self.canjes_tocados: set[int] = set()

    def g(self, fila, col):
        return self.hoja.cell(fila, ci(col)).value

    def _fechas_del_canje(self, fila, canje: Canje) -> dict:
        """La mejor fecha real por lado, con respaldo en cascada."""
        respaldo = (
            _fecha(self.g(fila, COL_FECHA_UPDATE))
            or canje.fecha_solicitud
        )
        return {
            "solicitante": _fecha(self.g(fila, COL_GESTION_SOLICITANTE)) or respaldo,
            "propietario": _fecha(self.g(fila, COL_GESTION_PROPIETARIO)) or respaldo,
            "acuerdo": (
                _fecha(self.g(fila, COL_ACUERDO_SOLICITANTE))
                or _fecha(self.g(fila, COL_ACUERDO_PROPIETARIO))
                or respaldo
            ),
            "general": respaldo,
        }

    def _agregar(self, canje_id: int, codigo: str, fecha, comentario: str) -> None:
        tipo = self.tipos.get(codigo)
        if tipo is None:
            raise RuntimeError(f"No existe el tipo de movimiento '{codigo}'")
        self.db.add(
            Movimiento(
                entity_type=EntityType.canje,
                entity_id=canje_id,
                tipo_movimiento=codigo,
                etapa_resultante=None,  # la migracion no mueve etapas: ya vienen de Dataprop
                fecha=fecha,
                autor_id=self.autor_id,
                comentario=comentario,
            )
        )
        self.creados[codigo] += 1
        self.canjes_tocados.add(canje_id)

    def procesar_fila(self, fila: int) -> None:
        bruto = self.g(fila, COL_ID)
        if bruto in (None, ""):
            return
        try:
            canje_id = int(float(bruto))
        except (TypeError, ValueError):
            return

        if canje_id not in self.canjes:
            self.sin_canje.append(canje_id)
            return

        canje = self.db.get(Canje, canje_id)
        fechas = self._fechas_del_canje(fila, canje)
        if _fecha(self.g(fila, COL_FECHA_UPDATE)) is None:
            self.sin_fecha.append(canje_id)

        operador = _texto(self.g(fila, COL_OPERADOR))
        firma = f"{NOTA}{f' · operador {operador}' if operador else ''}"

        # Los pasos completados, cada uno como su propio movimiento.
        for col, (codigo, lado) in CHECKLIST.items():
            if _es_si(self.g(fila, col)):
                self._agregar(canje_id, codigo, fechas[lado], firma)

        # Los pasos marcados como "No" y las observaciones no son pasos
        # completados, asi que van juntos en un comentario en vez de inventar
        # tipos de movimiento para representarlos.
        negativos = [
            self.tipos[CHECKLIST[c][0]].nombre
            for c in CHECKLIST
            if _es_no(self.g(fila, c))
        ]
        observaciones = _texto(self.g(fila, COL_OBSERVACIONES))
        partes = []
        if negativos:
            partes.append("No aplicaba o no se logró: " + ", ".join(negativos))
        if observaciones:
            partes.append(observaciones)
        if partes:
            self._agregar(
                canje_id, "COMENTARIO_GENERAL", fechas["general"],
                f"{firma}. " + " | ".join(partes),
            )

        # Cancelacion, solo cuando hay un tipo real.
        tipo_cancelacion = _texto(self.g(fila, COL_TIPO_CANCELACION))
        motivo = _texto(self.g(fila, COL_MOTIVO_CANCELACION))
        if tipo_cancelacion and tipo_cancelacion.upper() not in SIN_TIPO:
            detalle = f"{tipo_cancelacion}" + (f". {motivo}" if motivo else "")
            self._agregar(canje_id, "CANCELACION", fechas["general"], f"{firma}. {detalle}")
        elif motivo:
            # Hay motivo sin tipo: igual es informacion que no se quiere perder.
            self._agregar(
                canje_id, "COMENTARIO_GENERAL", fechas["general"],
                f"{firma}. Motivo de cancelación: {motivo}",
            )

    def correr(self) -> None:
        for fila in range(FILA_PRIMERA, self.hoja.max_row + 1):
            self.procesar_fila(fila)
        self.db.commit()


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(1)
    ruta = sys.argv[1]
    reemplazar = "--reemplazar" in sys.argv

    hoja = openpyxl.load_workbook(ruta, data_only=True)[HOJA]

    with SessionLocal() as db:
        ya_hay = db.scalar(
            select(Movimiento.id).where(Movimiento.entity_type == EntityType.canje).limit(1)
        )
        if ya_hay is not None:
            if not reemplazar:
                print(
                    "Ya hay movimientos de canjes registrados.\n"
                    "Usar --reemplazar para borrarlos y volver a migrar."
                )
                raise SystemExit(1)
            borrados = db.execute(
                delete(Movimiento).where(Movimiento.entity_type == EntityType.canje)
            ).rowcount
            db.commit()
            print(f"Borrados {borrados} movimientos de canjes anteriores.")

        autor = db.scalar(select(Usuario).where(Usuario.rol == "admin"))
        if autor is None:
            print("No hay un usuario admin al que atribuir la migración.")
            raise SystemExit(1)

        print(f"Migrando en {engine.url.host}, atribuido a {autor.nombre}\n")
        m = Migracion(db, hoja, autor.id)
        m.correr()

        print(f"{'tipo de movimiento':<30}creados")
        for codigo, n in sorted(m.creados.items(), key=lambda kv: -kv[1]):
            print(f"  {codigo:<28}{n:>5}")
        print(f"  {'TOTAL':<28}{sum(m.creados.values()):>5}")
        print()
        print(f"canjes con seguimiento migrado : {len(m.canjes_tocados)}")
        if m.sin_canje:
            print(f"filas cuyo canje no está en la base: {len(m.sin_canje)} -> {sorted(set(m.sin_canje))[:12]}")
        if m.sin_fecha:
            print(
                f"canjes sin 'Fecha último update': {len(set(m.sin_fecha))} "
                "(se usó la fecha de gestión o la de solicitud)"
            )


if __name__ == "__main__":
    main()
