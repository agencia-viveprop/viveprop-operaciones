"""Aplica el archivo de revisión de monedas: corrige `canjes.moneda_valor`.

Es el paso que `revisar_monedas_canjes` deliberadamente no hace. Ese genera la
lista; este escribe, y solo después de que una persona revisó el archivo.

**Solo toca la moneda.** No el valor, no la etapa, no el estado, nada más. La
corrección es de una etiqueta equivocada --un arriendo de casa en Vitacura
guardado como "70 CLP"-- y el monto que hay al lado es correcto: lo que estaba mal
era decir en qué unidad estaba expresado.

**No escribe salvo que se lo pidan.** Sin `--aplicar` hace una pasada en seco y
dice qué haría. Es el default porque esto corre contra producción, y una escritura
de 112 filas no puede ser el resultado de tipear mal un comando.

**Compara contra el estado actual antes de escribir.** Si un canje cambió en la
base después de que se generó el archivo, la revisión que trae ese archivo está
vieja y aplicarla pisaría una edición más nueva. Esas filas se omiten y se
informan, en vez de ganar por ser las últimas en llegar.

**Las filas sin decisión no se tocan.** Una celda vacía en «Moneda correcta»
significa "no revisada", no "borrala".

Se corre desde `backend/`:

    python -m app.scripts.aplicar_monedas_canjes                    # en seco
    python -m app.scripts.aplicar_monedas_canjes --aplicar          # escribe
"""
import os
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

RAIZ = Path(__file__).resolve().parent.parent.parent
ARCHIVO = RAIZ.parent / "Archivos" / "revision-monedas-canjes.xlsx"
HOJA = "REVISAR"

# Las columnas que se leen, por posición. El archivo lo genera el script hermano,
# así que el orden es conocido; se valida el encabezado de todas formas.
COL_CANJE = 1
COL_MONEDA_ARCHIVO = 8
COL_MONEDA_CORRECTA = 9
ENCABEZADOS = {COL_CANJE: "Canje", COL_MONEDA_ARCHIVO: "Moneda actual", COL_MONEDA_CORRECTA: "Moneda correcta"}

MONEDAS_VALIDAS = ("CLP", "UF", "OTRA")


@dataclass(frozen=True)
class Decision:
    canje_id: int
    # La que el archivo dice que tenía cuando se generó.
    moneda_en_el_archivo: str | None
    # La que la persona dejó en la columna editable.
    moneda_correcta: str


@dataclass
class Plan:
    cambios: list[tuple[int, str, str]]  # canje, de, a
    sin_decision: int = 0
    ya_estaban: int = 0
    desactualizadas: list[str] = None
    invalidas: list[str] = None
    inexistentes: list[str] = None

    def __post_init__(self):
        self.desactualizadas = self.desactualizadas or []
        self.invalidas = self.invalidas or []
        self.inexistentes = self.inexistentes or []


def planificar(decisiones: list[Decision], actual: dict[int, str | None]) -> Plan:
    """Qué cambiar, sin tocar la base. Es la parte que se puede probar.

    `actual` es la moneda que cada canje tiene **ahora**. Se compara contra la que
    el archivo dice que tenía: si no coinciden, alguien editó ese canje después de
    generar el archivo y esta revisión ya no aplica.
    """
    plan = Plan(cambios=[])
    for d in decisiones:
        if d.canje_id not in actual:
            plan.inexistentes.append(f"canje {d.canje_id}: no existe")
            continue
        if not d.moneda_correcta:
            plan.sin_decision += 1
            continue
        if d.moneda_correcta not in MONEDAS_VALIDAS:
            plan.invalidas.append(
                f"canje {d.canje_id}: '{d.moneda_correcta}' no es una moneda "
                f"({', '.join(MONEDAS_VALIDAS)})"
            )
            continue

        hoy = actual[d.canje_id]
        if hoy != d.moneda_en_el_archivo:
            plan.desactualizadas.append(
                f"canje {d.canje_id}: el archivo dice que tenía "
                f"{d.moneda_en_el_archivo or 'nada'} y ahora tiene {hoy or 'nada'}; "
                "alguien lo editó después. No se toca."
            )
            continue
        if hoy == d.moneda_correcta:
            plan.ya_estaban += 1
            continue

        plan.cambios.append((d.canje_id, hoy or "nada", d.moneda_correcta))
    return plan


def leer_archivo(ruta: Path) -> list[Decision]:
    libro = openpyxl.load_workbook(ruta, data_only=True)
    if HOJA not in libro.sheetnames:
        raise SystemExit(f"El archivo no tiene la hoja '{HOJA}'.")
    hoja = libro[HOJA]

    for col, esperado in ENCABEZADOS.items():
        leido = hoja.cell(row=1, column=col).value
        if (leido or "").strip() != esperado:
            raise SystemExit(
                f"La columna {col} dice '{leido}' y se esperaba '{esperado}'. "
                "¿Se movieron las columnas? Volvé a generar el archivo."
            )

    decisiones = []
    for fila in range(2, hoja.max_row + 1):
        canje = hoja.cell(row=fila, column=COL_CANJE).value
        if canje is None:
            continue
        decisiones.append(
            Decision(
                canje_id=int(canje),
                moneda_en_el_archivo=(hoja.cell(row=fila, column=COL_MONEDA_ARCHIVO).value or None),
                moneda_correcta=str(hoja.cell(row=fila, column=COL_MONEDA_CORRECTA).value or "").strip().upper(),
            )
        )
    return decisiones


def main() -> None:
    aplicar = "--aplicar" in sys.argv
    if not os.environ.get("DATABASE_URL"):
        load_dotenv(RAIZ / ".env")

    destino = os.environ["DATABASE_URL"].split("@")[-1].split("/")[0]
    print(f"archivo : {ARCHIVO}")
    print(f"base    : {destino}")
    print(f"modo    : {'APLICAR (escribe)' if aplicar else 'en seco (no escribe)'}")
    print()

    decisiones = leer_archivo(ARCHIVO)
    engine = create_engine(os.environ["DATABASE_URL"])
    with engine.begin() as c:
        actual = {
            id_: moneda
            for id_, moneda in c.execute(text("select id, moneda_valor::text from canjes")).all()
        }

        plan = planificar(decisiones, actual)

        print(f"filas en el archivo    : {len(decisiones)}")
        print(f"  a cambiar            : {len(plan.cambios)}")
        print(f"  ya estaban bien      : {plan.ya_estaban}")
        print(f"  sin decision (vacias): {plan.sin_decision}")
        for titulo, lista in (
            ("desactualizadas", plan.desactualizadas),
            ("moneda invalida", plan.invalidas),
            ("no existen", plan.inexistentes),
        ):
            if lista:
                print(f"  {titulo:21}: {len(lista)}")
                for x in lista:
                    print(f"      {x}")
        print()

        for canje_id, de, a in plan.cambios:
            print(f"  canje {canje_id:<5} {de:>4} -> {a}")

        if not aplicar:
            print()
            print("Nada se escribió. Para aplicar: --aplicar")
            return

        for canje_id, _, a in plan.cambios:
            c.execute(
                text("update canjes set moneda_valor = :m where id = :id"),
                {"m": a, "id": canje_id},
            )
        print()
        print(f"{len(plan.cambios)} canjes actualizados.")


if __name__ == "__main__":
    main()
