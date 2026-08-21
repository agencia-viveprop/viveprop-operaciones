"""Plantilla y carga manual de la serie de UF (sprint 5, ver D-007).

No hay integración con ninguna fuente externa: se descarga una plantilla, se
pega la serie del mes y se sube. La UF se publica del día 10 al 9 del mes
siguiente, así que siempre hay un tramo conocido por delante y esto es una tarea
de una vez al mes.

**Este módulo es el piloto del patrón** que reusan los sprints 14 y 15 para la
carga masiva de negocios: generar plantilla, validar fila por fila, informar qué
pasó con cada una, y hacer upsert idempotente.

La plantilla viene **con las fechas que faltan ya escritas**, para que no haya
que averiguar cuáles son: se rellenan los valores y se sube.
"""
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.models.uf import UFDiaria
from app.services.uf import dias_de_colchon, rango

COL_FECHA, COL_VALOR = "FECHA", "VALOR"
# Cuantos dias hacia adelante trae la plantilla. Cubre el tramo del mes que
# viene con holgura, sin volverse una lista interminable.
DIAS_PLANTILLA = 45
# Umbral del aviso, en dias de serie restante. Ver D-008: con 3 dias se enciende
# entre el 6 y el 9, cuando la publicacion del nuevo tramo ya salio. Con 15 se
# quedaria prendido media vida y dejaria de ser un aviso.
UMBRAL_AVISO = 3


class EstadoSerie(BaseModel):
    primera: date | None
    ultima: date | None
    filas: int
    dias_de_colchon: int | None
    # 'vacia' | 'vencida' | 'aviso' | 'ok'
    nivel: str
    mensaje: str


class ResumenCargaUF(BaseModel):
    nuevas: int = 0
    actualizadas: int = 0
    sin_cambio: int = 0
    errores: list[str] = []

    @property
    def hubo_cambios(self) -> bool:
        return bool(self.nuevas or self.actualizadas)


@dataclass
class _Fila:
    fecha: date
    valor: Decimal


def estado_serie(db: Session, hoy: date) -> EstadoSerie:
    """Qué tan lejos llega la serie, y si hay que hacer algo al respecto."""
    primera, ultima = rango(db)
    if ultima is None:
        return EstadoSerie(
            primera=None, ultima=None, filas=0, dias_de_colchon=None, nivel="vacia",
            mensaje="La serie de UF está vacía. Hay que cargarla para poder valorizar negocios.",
        )

    filas = db.query(UFDiaria).count()
    colchon = dias_de_colchon(db, hoy)

    if colchon < 0:
        nivel = "vencida"
        mensaje = (
            f"La serie de UF venció el {ultima.isoformat()}. "
            "No se pueden valorizar negocios con fecha de hoy hasta cargar el nuevo tramo."
        )
    elif colchon <= UMBRAL_AVISO:
        nivel = "aviso"
        dias = "hoy" if colchon == 0 else f"en {colchon} día{'s' if colchon != 1 else ''}"
        mensaje = f"La serie de UF llega hasta el {ultima.isoformat()}: se agota {dias}."
    else:
        nivel = "ok"
        mensaje = f"La serie de UF llega hasta el {ultima.isoformat()}, {colchon} días por delante."

    return EstadoSerie(
        primera=primera, ultima=ultima, filas=filas, dias_de_colchon=colchon,
        nivel=nivel, mensaje=mensaje,
    )


def generar_plantilla(db: Session, hoy: date) -> bytes:
    """Un .xlsx con las fechas que faltan y la columna de valor en blanco.

    Si la serie está vacía, la plantilla arranca en el día de hoy: no tiene
    sentido pedir cuatro años de historia a mano.
    """
    _, ultima = rango(db)
    desde = (ultima + timedelta(days=1)) if ultima is not None else hoy

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "UF"

    encabezado = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="3D3EA8")
    for col, titulo in enumerate([COL_FECHA, COL_VALOR], start=1):
        celda = hoja.cell(1, col, titulo)
        celda.font = encabezado
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center")
    hoja.column_dimensions["A"].width = 14
    hoja.column_dimensions["B"].width = 14

    for i in range(DIAS_PLANTILLA):
        fila = i + 2
        hoja.cell(fila, 1, (desde + timedelta(days=i)).isoformat())
        hoja.cell(fila, 2, None).number_format = "#,##0.00"

    guia = libro.create_sheet("Instrucciones")
    for i, linea in enumerate(
        [
            "Cómo cargar la UF",
            "",
            "1. En la hoja 'UF', rellena la columna VALOR con el valor de cada fecha.",
            "2. Las filas con VALOR vacío se ignoran: no hace falta borrarlas.",
            "3. Sube el archivo desde la app, en Mantención → UF.",
            "",
            "La carga es por fecha, así que subir un archivo que se solapa con lo",
            "que ya está cargado no duplica nada: actualiza los valores que cambien",
            "y deja igual los que no.",
            "",
            "La UF se publica del día 10 de cada mes al 9 del siguiente, así que",
            "basta hacer esto una vez al mes.",
        ],
        start=1,
    ):
        celda = guia.cell(i, 1, linea)
        if i == 1:
            celda.font = Font(bold=True, size=13)
    guia.column_dimensions["A"].width = 80

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _parsear_fecha(valor) -> date:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return date.fromisoformat(str(valor).strip()[:10])


def _parsear_valor(valor) -> Decimal:
    if isinstance(valor, (int, float, Decimal)):
        numero = Decimal(str(valor))
    else:
        # Se aceptan las dos convenciones: "40.885,63" y "40885.63".
        texto = str(valor).strip().replace("$", "").replace(" ", "")
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        numero = Decimal(texto)
    if numero <= 0:
        raise ValueError("el valor de la UF tiene que ser positivo")
    return numero.quantize(Decimal("0.01"))


def cargar_desde_xlsx(db: Session, contenido: bytes) -> ResumenCargaUF:
    """Lee la plantilla y hace upsert por fecha, informando fila por fila.

    Si hay errores de formato **no se carga nada**: es preferible que la persona
    corrija el archivo antes que quedar con media serie subida y no saber cuál
    mitad.
    """
    try:
        libro = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo: {exc}") from exc

    hoja = libro["UF"] if "UF" in libro.sheetnames else libro.worksheets[0]

    encabezados = [
        str(c.value).strip().upper() if c.value is not None else "" for c in hoja[1]
    ]
    faltantes = [c for c in (COL_FECHA, COL_VALOR) if c not in encabezados]
    if faltantes:
        raise ValueError(
            f"Faltan columnas en el archivo: {', '.join(faltantes)}. "
            f"Se encontraron: {', '.join(e for e in encabezados if e) or '(ninguna)'}."
        )
    i_fecha, i_valor = encabezados.index(COL_FECHA), encabezados.index(COL_VALOR)

    resumen = ResumenCargaUF()
    filas: dict[date, Decimal] = {}

    for n in range(2, hoja.max_row + 1):
        celdas = [c.value for c in hoja[n]]
        bruto_fecha = celdas[i_fecha] if i_fecha < len(celdas) else None
        bruto_valor = celdas[i_valor] if i_valor < len(celdas) else None

        if bruto_fecha is None and bruto_valor is None:
            continue
        if bruto_valor is None or str(bruto_valor).strip() == "":
            # Fila de la plantilla que no se lleno: se ignora sin ruido.
            continue

        try:
            fecha = _parsear_fecha(bruto_fecha)
        except Exception:
            resumen.errores.append(f"Fila {n}: fecha inválida ({bruto_fecha!r})")
            continue
        try:
            valor = _parsear_valor(bruto_valor)
        except (InvalidOperation, ValueError, ArithmeticError) as exc:
            resumen.errores.append(f"Fila {n}: valor inválido ({bruto_valor!r}) — {exc}")
            continue

        if fecha in filas and filas[fecha] != valor:
            resumen.errores.append(
                f"Fila {n}: el {fecha.isoformat()} aparece dos veces con valores distintos"
            )
            continue
        filas[fecha] = valor

    if resumen.errores:
        return resumen
    if not filas:
        return resumen

    return guardar_serie(db, filas)


def guardar_serie(db: Session, filas: dict[date, Decimal]) -> ResumenCargaUF:
    """Upsert por fecha, informando cuantas son nuevas, cambiadas y iguales.

    Vive aparte porque tiene dos usuarios: la carga manual de la plantilla y la
    descarga automatica desde el SII. Que los dos caminos escriban con el mismo
    codigo es lo que hace que la automatizacion no pueda dejar la serie en un
    estado que la carga manual no produciria.

    Se usa el ORM y no `ON CONFLICT` de Postgres a proposito: son 45 filas al
    mes, el rendimiento es irrelevante, y asi este calculo se puede testear
    contra la base en memoria. El upsert lo da `merge`, que inserta o actualiza
    segun la clave primaria.
    """
    resumen = ResumenCargaUF()
    if not filas:
        return resumen

    existentes = {
        f.fecha: f.valor
        for f in db.query(UFDiaria).filter(UFDiaria.fecha.in_(list(filas))).all()
    }
    hubo = False
    for fecha, valor in sorted(filas.items()):
        anterior = existentes.get(fecha)
        if anterior is None:
            resumen.nuevas += 1
        elif Decimal(anterior) != valor:
            resumen.actualizadas += 1
        else:
            resumen.sin_cambio += 1
            continue
        db.merge(UFDiaria(fecha=fecha, valor=valor))
        hubo = True

    if hubo:
        db.commit()

    return resumen
