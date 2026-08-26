"""Carga del historial de etapas de negocios.

Lee la plantilla de `plantilla_historial` y crea los movimientos que hoy no
existen, para que la proyección de plazos pase de declararse imposible a
calcularse. Cuatro reglas la gobiernan, y las cuatro se acordaron antes de
escribirla.

**No agenda próxima acción.** Si lo hiciera, cargar la historia de 18 negocios
metería 18 compromisos vencidos de meses en «Qué me toca hoy», y la pantalla
quedaría en rojo por una tarea administrativa. Una carga histórica no es una
gestión: nadie prometió volver el jueves.

**No puede hacer retroceder la etapa vigente.** Registrar un movimiento mueve la
etapa del negocio a la del movimiento cronológicamente más nuevo (`D-060`). Cargar
`E1` y `E2` de un negocio que hoy está en `E7` lo bajaría a `E2` y borraría el
dato bueno con uno viejo. Acá la etapa del negocio **no se toca**: la carga
escribe historia, no cambia el presente.

**Recargar no duplica.** La clave es negocio + etapa: subir el archivo otra vez
con una fecha corregida actualiza ese movimiento en vez de agregar otro. Es lo que
permite iterar --cargar lo que se sabe, revisar el resultado, corregir-- sin
ensuciar la bitácora.

**La validación de fecha mínima se relaja, y por eso hay que relajarla.** El
sistema rechaza un movimiento anterior al inicio de la primera liquidación del
negocio, que es correcto para el uso normal. Pero en 7 liquidaciones esa fecha de
inicio **está mal** --es la de cierre, porque el Excel traía una sola-- así que la
validación bloquearía justamente las fechas reales que se quieren cargar. Se
permite y se informa cuáles quedaron antes del inicio registrado.

**La corrección de fechas de inicio no puede mover plata, y eso se comprueba, no
se supone.** Si una liquidación no tiene `fecha_valorizacion`, su UF sale de
`fecha_inicio`: cambiarla movería el monto y la comisión. La carga se niega a
tocar esas, y lo dice. Hoy ninguna de las siete está en ese caso --cinco tienen
`fecha_valorizacion` y la sexta y séptima tienen valor manual, que manda sobre la
conversión (`D-017`)-- pero la guarda es mecánica para que siga siendo verdad si
los datos cambian.
"""
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from io import BytesIO

import openpyxl
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalogo import Etapa
from app.models.movimiento import EntityType, Movimiento, TipoMovimiento
from app.models.negocio import Negocio, NegocioHito
from app.services.plantilla_historial import (
    COLUMNAS_HISTORIAL,
    COLUMNAS_LIQUIDACIONES,
    FILA_ENCABEZADO,
    HOJA_HISTORIAL,
    HOJA_LIQUIDACIONES,
)


class ImportarHistorialError(Exception):
    """El archivo no se puede leer. Distinto de una fila con problemas."""


class ResumenHistorial(BaseModel):
    movimientos_creados: int = 0
    movimientos_actualizados: int = 0
    filas_sin_fecha: int = 0
    fechas_corregidas: int = 0
    # Filas que no se pudieron aplicar, con el motivo. Van todas: un resumen que
    # muestra las primeras cinco esconde justo el problema que se repite.
    omitidas: list[str] = []
    # Movimientos cuya fecha quedó antes del inicio registrado de la liquidación.
    # No es un error --es el motivo por el que existe la hoja LIQUIDACIONES-- pero
    # hay que verlo para saber cuáles corregir.
    anteriores_al_inicio: list[str] = []
    # Liquidaciones que la carga se negó a corregir porque su plata depende de la
    # fecha de inicio.
    no_corregidas_por_plata: list[str] = []

    @property
    def total_movimientos(self) -> int:
        return self.movimientos_creados + self.movimientos_actualizados


@dataclass
class _Contexto:
    """Lo que hace falta resolver una vez y no por fila."""

    negocios: dict[str, Negocio] = field(default_factory=dict)
    etapas: dict[str, Etapa] = field(default_factory=dict)
    # El tipo de movimiento de cada etapa: el que la deja como resultado.
    tipo_de_etapa: dict[str, str] = field(default_factory=dict)


def _leer_fecha(valor) -> date | None:
    """Una fecha de celda, sea texto o fecha de Excel.

    Se aceptan los dos porque las dos aparecen: quien escribe a mano deja texto y
    quien copia de otra planilla deja una fecha de verdad. Rechazar una de las dos
    sería rechazar la mitad de los archivos por una razón invisible.
    """
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for formato in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    raise ValueError(f"no se entiende la fecha '{texto}'")


def _validar_encabezado(hoja, columnas, nombre_hoja: str) -> None:
    esperados = [c.nombre for c in columnas]
    leidos = [
        (hoja.cell(row=FILA_ENCABEZADO, column=i).value or "").strip() if isinstance(
            hoja.cell(row=FILA_ENCABEZADO, column=i).value, str
        ) else hoja.cell(row=FILA_ENCABEZADO, column=i).value
        for i in range(1, len(esperados) + 1)
    ]
    faltan = [e for e in esperados if e not in leidos]
    if faltan:
        raise ImportarHistorialError(
            f"La hoja {nombre_hoja} no tiene las columnas esperadas. Faltan: {', '.join(faltan)}. "
            "Bajá la plantilla de nuevo en vez de armar el archivo a mano."
        )


def _contexto(db: Session) -> _Contexto:
    ctx = _Contexto()
    for n in db.scalars(select(Negocio)).all():
        ctx.negocios[n.codigo.strip().upper()] = n
    for e in db.scalars(select(Etapa)).all():
        ctx.etapas[e.codigo.strip().upper()] = e
    # De todos los tipos que dejan una etapa, se toma el de menor orden: es el
    # paso canónico hacia esa etapa. Si hubiera varios, elegir el primero es
    # determinista, que es lo que importa para que recargar no cambie nada.
    for tipo in db.scalars(
        select(TipoMovimiento)
        .where(
            TipoMovimiento.entity_type == EntityType.negocio,
            TipoMovimiento.etapa_resultante.is_not(None),
        )
        .order_by(TipoMovimiento.orden, TipoMovimiento.codigo)
    ).all():
        ctx.tipo_de_etapa.setdefault(tipo.etapa_resultante, tipo.codigo)
    return ctx


def _cargar_historial(db: Session, hoja, ctx: _Contexto, resumen: ResumenHistorial, autor_id: int | None) -> None:
    for fila in range(FILA_ENCABEZADO + 1, hoja.max_row + 1):
        codigo = hoja.cell(row=fila, column=1).value
        etapa_codigo = hoja.cell(row=fila, column=2).value
        if not codigo and not etapa_codigo:
            continue

        donde = f"fila {fila}"
        codigo = str(codigo or "").strip().upper()
        etapa_codigo = str(etapa_codigo or "").strip().upper()

        try:
            cuando = _leer_fecha(hoja.cell(row=fila, column=3).value)
        except ValueError as exc:
            resumen.omitidas.append(f"{donde} ({codigo} {etapa_codigo}): {exc}")
            continue

        if cuando is None:
            resumen.filas_sin_fecha += 1
            continue

        negocio = ctx.negocios.get(codigo)
        if negocio is None:
            resumen.omitidas.append(f"{donde}: no existe el negocio '{codigo}'")
            continue
        if etapa_codigo not in ctx.etapas:
            resumen.omitidas.append(f"{donde} ({codigo}): no existe la etapa '{etapa_codigo}'")
            continue

        tipo = ctx.tipo_de_etapa.get(etapa_codigo)
        if tipo is None:
            resumen.omitidas.append(
                f"{donde} ({codigo}): ningún tipo de movimiento deja el negocio en '{etapa_codigo}'"
            )
            continue

        hoy = datetime.now(timezone.utc).date()
        if cuando > hoy:
            resumen.omitidas.append(
                f"{donde} ({codigo} {etapa_codigo}): la fecha está en el futuro"
            )
            continue

        descripcion = hoja.cell(row=fila, column=4).value
        comentario = str(descripcion).strip() if descripcion else None
        momento = datetime.combine(cuando, time.min, tzinfo=timezone.utc)

        # La clave es negocio + etapa: recargar corrige en vez de duplicar.
        existente = db.scalar(
            select(Movimiento).where(
                Movimiento.entity_type == EntityType.negocio,
                Movimiento.entity_id == negocio.id,
                Movimiento.etapa_resultante == etapa_codigo,
            )
        )
        if existente is not None:
            existente.fecha = momento
            existente.comentario = comentario
            existente.tipo_movimiento = tipo
            resumen.movimientos_actualizados += 1
        else:
            db.add(
                Movimiento(
                    entity_type=EntityType.negocio,
                    entity_id=negocio.id,
                    tipo_movimiento=tipo,
                    etapa_resultante=etapa_codigo,
                    fecha=momento,
                    autor_id=autor_id,
                    comentario=comentario,
                    # Sin próxima acción: ver el docstring del módulo.
                    proximo_seguimiento=None,
                )
            )
            resumen.movimientos_creados += 1

        inicio = min((h.fecha_inicio for h in negocio.hitos if h.fecha_inicio), default=None)
        if inicio is not None and cuando < inicio:
            resumen.anteriores_al_inicio.append(
                f"{codigo} {etapa_codigo}: {cuando:%d-%m-%Y}, antes del inicio "
                f"registrado ({inicio:%d-%m-%Y})"
            )


def _corregir_liquidaciones(db: Session, hoja, ctx: _Contexto, resumen: ResumenHistorial) -> None:
    for fila in range(FILA_ENCABEZADO + 1, hoja.max_row + 1):
        codigo = hoja.cell(row=fila, column=1).value
        if not codigo:
            continue
        codigo = str(codigo).strip().upper()
        nombre = hoja.cell(row=fila, column=2).value
        nombre = str(nombre).strip() if nombre else None
        donde = f"fila {fila} de {HOJA_LIQUIDACIONES}"

        try:
            inicio_real = _leer_fecha(hoja.cell(row=fila, column=3).value)
        except ValueError as exc:
            resumen.omitidas.append(f"{donde} ({codigo}): {exc}")
            continue
        if inicio_real is None:
            continue

        negocio = ctx.negocios.get(codigo)
        if negocio is None:
            resumen.omitidas.append(f"{donde}: no existe el negocio '{codigo}'")
            continue

        # "ÚNICA" es lo que la plantilla escribe cuando la liquidación no tiene
        # nombre; hay que traducirlo de vuelta a nulo para encontrarla.
        buscado = None if nombre in (None, "ÚNICA", "UNICA") else nombre
        candidatos = [h for h in negocio.hitos if (h.nombre or None) == buscado]
        if len(candidatos) != 1:
            resumen.omitidas.append(
                f"{donde} ({codigo}): no se pudo identificar la liquidación '{nombre}'"
            )
            continue

        hito = candidatos[0]
        if inicio_real > (hito.fecha_cierre or inicio_real):
            resumen.omitidas.append(
                f"{donde} ({codigo}): el inicio real es posterior al cierre registrado"
            )
            continue

        # La guarda: si la plata de esta liquidación sale de `fecha_inicio`,
        # cambiarla movería el monto. Se comprueba, no se supone.
        if hito.fecha_valorizacion is None and hito.valor_clp_manual is None:
            resumen.no_corregidas_por_plata.append(
                f"{codigo} {nombre or 'única'}: su valorización depende de la fecha de inicio, "
                "así que corregirla movería el monto y la comisión"
            )
            continue

        hito.fecha_inicio = inicio_real
        resumen.fechas_corregidas += 1


def importar_historial(
    db: Session, contenido: bytes, autor_id: int | None = None
) -> ResumenHistorial:
    """Carga el historial y corrige las fechas de inicio, en una sola transacción.

    Todo o nada: si algo revienta, no queda media historia cargada. Las filas con
    problemas no revientan --se omiten y se informan-- así que un archivo con tres
    errores carga las demás y dice cuáles no.
    """
    try:
        libro = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl lanza de todo
        raise ImportarHistorialError(f"No se pudo leer el archivo: {exc}") from exc

    if HOJA_HISTORIAL not in libro.sheetnames:
        raise ImportarHistorialError(
            f"El archivo no tiene la hoja '{HOJA_HISTORIAL}'. Bajá la plantilla de nuevo."
        )

    resumen = ResumenHistorial()
    ctx = _contexto(db)

    historial = libro[HOJA_HISTORIAL]
    _validar_encabezado(historial, COLUMNAS_HISTORIAL, HOJA_HISTORIAL)

    # La etapa vigente de cada negocio, para restaurarla después: cargar
    # movimientos la movería, y la carga escribe historia sin cambiar el presente.
    etapas_antes = {n.id: n.etapa for n in ctx.negocios.values()}

    _cargar_historial(db, historial, ctx, resumen, autor_id)

    if HOJA_LIQUIDACIONES in libro.sheetnames:
        liquidaciones = libro[HOJA_LIQUIDACIONES]
        _validar_encabezado(liquidaciones, COLUMNAS_LIQUIDACIONES, HOJA_LIQUIDACIONES)
        _corregir_liquidaciones(db, liquidaciones, ctx, resumen)

    db.flush()
    for negocio in ctx.negocios.values():
        negocio.etapa = etapas_antes[negocio.id]

    db.commit()
    return resumen
