"""Carga masiva de negocios desde la plantilla.

**Si hay un solo error no se carga nada.** Se revisan todas las filas, se
informan todos los problemas juntos y la base queda intacta. Es la misma regla
de la carga de UF, por el mismo motivo: media carga es peor que ninguna, porque
nadie sabe cuál mitad quedó y hay que revisar fila por fila para averiguarlo.

**Cargar dos veces no duplica.** Un código que ya existe se actualiza, y un hito
se reconoce por su negocio y su nombre. Eso es lo que permite corregir el archivo
y volver a subirlo, que es como la gente realmente trabaja.

**Nunca borra.** Si la base tiene dos hitos y el archivo trae uno, el otro se
queda. Un import que borra lo que no menciona convierte un archivo incompleto en
pérdida de datos.

**El motor calcula las comisiones, siempre.** La plantilla pide el valor y las
tasas; comisión total, broker, rebate, VP bruta, equipo, tercero y real VP salen
de `refrescar_hito`. Por eso esta **no** es la herramienta para los 19
históricos: esos se migran fieles y sin recalcular (`D-026`) con
`scripts/cargar_negocios.py`, porque siete están cerrados con plata ya facturada.
"""
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalogo import Catalogo, EstadoNegocio, Etapa, ModeloNegocio
from app.models.negocio import MonedaTipo, Negocio, NegocioHito
from app.services.negocios import NegocioError, obtener_o_crear_propiedad, refrescar_hito
from app.services.plantilla_negocios import HOJA, NOMBRES, OBLIGATORIAS

# El encabezado va en la fila 2: la 1 lleva los nombres de grupo.
FILA_ENCABEZADO = 2
FILA_DATOS = FILA_ENCABEZADO + 1

CIEN = Decimal("100")

COLUMNAS_PCT = (
    "PCT_LADO_VENDEDOR", "PCT_LADO_COMPRADOR", "PCT_BROKER_VENDEDOR",
    "PCT_BROKER_COMPRADOR", "PCT_VP_VENDEDOR", "PCT_VP_COMPRADOR",
    "PCT_REBATE_CONCENTRADOR", "PCT_EQUIPO", "PCT_TERCERO",
)

# Los datos de nivel negocio: si dos filas comparten CODIGO tienen que coincidir
# en todos estos, porque no hay forma de saber cuál gana.
COLUMNAS_NEGOCIO = (
    "DIRECCION", "UNIDAD", "COMUNA", "TIPO_PROPIEDAD", "MODELO", "ALIANZA",
    "TIPO_OPERACION", "ETAPA", "VENDEDOR_ARRENDADOR", "COMPRADOR_ARRENDATARIO",
    "CORREDOR_AGENTE",
)


class ResumenCargaNegocios(BaseModel):
    negocios_nuevos: int = 0
    negocios_actualizados: int = 0
    hitos_nuevos: int = 0
    hitos_actualizados: int = 0
    errores: list[str] = []

    @property
    def hubo_cambios(self) -> bool:
        return bool(
            self.negocios_nuevos or self.negocios_actualizados
            or self.hitos_nuevos or self.hitos_actualizados
        )


@dataclass
class _Fila:
    numero: int
    valores: dict[str, object] = field(default_factory=dict)


class ArchivoInvalido(Exception):
    """El archivo no se pudo leer, o no tiene la forma de la plantilla."""


# ------------------------------------------------------------- conversiones


def _texto(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _fecha(valor, campo: str) -> date | None:
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()[:10]
    for formato in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    raise ValueError(f"{campo}: fecha inválida ({valor!r}). Usá 2026-08-21 o 21-08-2026.")


def _decimal(valor, campo: str) -> Decimal | None:
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    texto = str(valor).strip().replace("$", "").replace(" ", "")
    # Las dos convenciones: 1.234,56 y 1234.56. Si hay coma, manda como decimal.
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ValueError(f"{campo}: número inválido ({valor!r}).") from exc


def _porcentaje(valor, campo: str) -> Decimal | None:
    """La plantilla pide 2 para 2%; la base guarda la fracción 0,02.

    Se pide en porcentaje porque es como está escrito en los contratos y en la
    hoja de reglas. Convertir acá es una división; pedirle a alguien que escriba
    0,0252001208200461 es pedirle que se equivoque.
    """
    bruto = _decimal(valor, campo)
    return None if bruto is None else bruto / CIEN


def _enum(valor, opciones, campo: str) -> str | None:
    texto = _texto(valor)
    if texto is None:
        return None
    arriba = texto.upper().replace(" ", "_")
    validos = {o.value for o in opciones}
    if arriba not in validos:
        raise ValueError(
            f"{campo}: '{texto}' no es un valor válido. Los válidos son: {', '.join(sorted(validos))}."
        )
    return arriba


# ------------------------------------------------------------------ lectura


def _leer_filas(contenido: bytes) -> list[_Fila]:
    try:
        libro = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    except Exception as exc:
        raise ArchivoInvalido(f"No se pudo leer el archivo: {exc}") from exc

    hoja = libro[HOJA] if HOJA in libro.sheetnames else libro.worksheets[0]
    encabezados = [
        _texto(c.value).upper() if _texto(c.value) else ""
        for c in hoja[FILA_ENCABEZADO]
    ]

    faltantes = [c for c in OBLIGATORIAS if c not in encabezados]
    if faltantes:
        raise ArchivoInvalido(
            f"Al archivo le faltan columnas obligatorias: {', '.join(faltantes)}. "
            f"Descargá la plantilla de nuevo. Se encontraron: "
            f"{', '.join(e for e in encabezados if e) or '(ninguna)'}."
        )

    indices = {nombre: i for i, nombre in enumerate(encabezados) if nombre in NOMBRES}

    filas: list[_Fila] = []
    for n in range(FILA_DATOS, hoja.max_row + 1):
        celdas = [c.value for c in hoja[n]]
        valores = {
            nombre: (celdas[i] if i < len(celdas) else None)
            for nombre, i in indices.items()
        }
        if all(v is None or str(v).strip() == "" for v in valores.values()):
            continue  # fila vacía de la plantilla
        filas.append(_Fila(numero=n, valores=valores))
    return filas


def _validar_fila(fila: _Fila, catalogos: dict, etapas: set[str]) -> dict:
    """Convierte y valida una fila.

    Devuelve los valores limpios, o un dict con `_errores` si hubo problemas.
    Acumula todos los errores de la fila en vez de cortar en el primero: quien
    corrige el archivo prefiere ver los tres problemas de una fila juntos.
    """
    errores: list[str] = []
    v = fila.valores
    limpio: dict = {}

    for col in OBLIGATORIAS:
        if _texto(v.get(col)) is None:
            errores.append(f"{col} no puede quedar vacío")

    def intentar(nombre, funcion, *args):
        try:
            limpio[nombre] = funcion(*args)
        except ValueError as exc:
            errores.append(str(exc))

    limpio["CODIGO"] = _texto(v.get("CODIGO"))
    limpio["HITO"] = _texto(v.get("HITO"))
    for col in ("DIRECCION", "UNIDAD", "COMUNA", "VENDEDOR_ARRENDADOR",
                "COMPRADOR_ARRENDATARIO", "CORREDOR_AGENTE", "NOTAS",
                "NOMBRE_TERCERO", "MOTIVO_VALOR_MANUAL"):
        limpio[col] = _texto(v.get(col))

    intentar("MODELO", _enum, v.get("MODELO"), ModeloNegocio, "MODELO")
    intentar("ESTADO", _enum, v.get("ESTADO"), EstadoNegocio, "ESTADO")
    intentar("MONEDA", _enum, v.get("MONEDA"), MonedaTipo, "MONEDA")
    intentar("FECHA_INICIO", _fecha, v.get("FECHA_INICIO"), "FECHA_INICIO")
    intentar("FECHA_CIERRE", _fecha, v.get("FECHA_CIERRE"), "FECHA_CIERRE")
    intentar("FECHA_VALORIZACION", _fecha, v.get("FECHA_VALORIZACION"), "FECHA_VALORIZACION")
    intentar("VALOR_NEGOCIO", _decimal, v.get("VALOR_NEGOCIO"), "VALOR_NEGOCIO")
    intentar("VALOR_CLP_MANUAL", _decimal, v.get("VALOR_CLP_MANUAL"), "VALOR_CLP_MANUAL")
    for col in COLUMNAS_PCT:
        intentar(col, _porcentaje, v.get(col), col)

    etapa = _texto(v.get("ETAPA"))
    if etapa is not None:
        etapa = etapa.upper()
        if etapa not in etapas:
            errores.append(f"ETAPA: '{etapa}' no existe. Las válidas son: {', '.join(sorted(etapas))}.")
    limpio["ETAPA"] = etapa

    for col, tipo in (("ALIANZA", "alianza"), ("TIPO_OPERACION", "tipo_operacion"),
                      ("TIPO_PROPIEDAD", "tipo_propiedad")):
        codigo = _texto(v.get(col))
        if codigo is None:
            limpio[col] = None
            continue
        arriba = codigo.upper()
        if arriba not in catalogos[tipo]:
            errores.append(
                f"{col}: '{codigo}' no existe. Mirá la hoja «Valores válidos» de la plantilla."
            )
            limpio[col] = None
        else:
            limpio[col] = arriba

    # Coherencias que ninguna columna sola puede ver.
    if limpio.get("ESTADO") == EstadoNegocio.CERRADO.value and limpio.get("FECHA_CIERRE") is None:
        errores.append("ESTADO es CERRADO pero FECHA_CIERRE está vacía")
    if limpio.get("FECHA_CIERRE") and limpio.get("FECHA_INICIO") \
            and limpio["FECHA_CIERRE"] < limpio["FECHA_INICIO"]:
        errores.append("FECHA_CIERRE es anterior a FECHA_INICIO")
    if limpio.get("VALOR_NEGOCIO") is not None and limpio.get("MONEDA") is None:
        errores.append("VALOR_NEGOCIO tiene monto pero MONEDA está vacía")
    if limpio.get("MONEDA") == MonedaTipo.UF.value and limpio.get("FECHA_VALORIZACION") is None:
        errores.append("MONEDA es UF pero FECHA_VALORIZACION está vacía: sin fecha no hay conversión")

    if errores:
        return {"_fila": fila.numero, "_errores": errores}
    return limpio | {"_fila": fila.numero}


# ------------------------------------------------------------------- carga


def cargar_desde_xlsx(db: Session, contenido: bytes) -> ResumenCargaNegocios:
    """Lee la plantilla, valida todo y recién entonces escribe."""
    filas = _leer_filas(contenido)
    resumen = ResumenCargaNegocios()
    if not filas:
        resumen.errores.append("El archivo no tiene ninguna fila con datos.")
        return resumen

    catalogos = {
        tipo: {
            c.codigo.upper(): c.id
            for c in db.execute(
                select(Catalogo).where(Catalogo.tipo == tipo, Catalogo.activo.is_(True))
            ).scalars()
        }
        for tipo in ("alianza", "tipo_operacion", "tipo_propiedad")
    }
    etapas = set(db.execute(select(Etapa.codigo)).scalars())

    limpias: list[dict] = []
    for fila in filas:
        resultado = _validar_fila(fila, catalogos, etapas)
        if "_errores" in resultado:
            for e in resultado["_errores"]:
                resumen.errores.append(f"Fila {resultado['_fila']}: {e}")
        else:
            limpias.append(resultado)

    # Un hito se identifica por su negocio y su nombre: el mismo par dos veces en
    # el archivo no se puede resolver.
    vistos: set[tuple[str, str]] = set()
    for f in limpias:
        clave = (f["CODIGO"], f["HITO"] or "")
        if clave in vistos:
            cual = f"'{f['HITO']}'" if f["HITO"] else "sin nombre"
            resumen.errores.append(
                f"Fila {f['_fila']}: el hito {cual} de {f['CODIGO']} aparece dos veces en el archivo"
            )
        vistos.add(clave)

    # Los datos de nivel negocio tienen que coincidir entre las filas del mismo código.
    por_codigo: dict[str, list[dict]] = {}
    for f in limpias:
        por_codigo.setdefault(f["CODIGO"], []).append(f)
    for codigo, grupo in por_codigo.items():
        primera = grupo[0]
        for otra in grupo[1:]:
            distintas = [c for c in COLUMNAS_NEGOCIO if primera.get(c) != otra.get(c)]
            if distintas:
                resumen.errores.append(
                    f"Fila {otra['_fila']}: {codigo} ya aparece en la fila {primera['_fila']} "
                    f"con otro valor en {', '.join(distintas)}. Los datos del negocio tienen que "
                    "ser iguales en todas sus filas."
                )

    if resumen.errores:
        return resumen

    try:
        _escribir(db, por_codigo, catalogos, resumen)
    except NegocioError as exc:
        db.rollback()
        resumen.errores.append(str(exc))
        return ResumenCargaNegocios(errores=resumen.errores)

    db.commit()
    return resumen


def _escribir(db: Session, por_codigo: dict, catalogos: dict, resumen: ResumenCargaNegocios) -> None:
    for codigo, grupo in por_codigo.items():
        cabeza = grupo[0]
        negocio = db.scalar(select(Negocio).where(Negocio.codigo == codigo))

        propiedad = obtener_o_crear_propiedad(db, {
            "direccion": cabeza["DIRECCION"],
            "unidad": cabeza["UNIDAD"],
            "comuna": cabeza["COMUNA"],
            "tipo_propiedad_id": catalogos["tipo_propiedad"].get(cabeza["TIPO_PROPIEDAD"] or ""),
        })

        campos_negocio = {
            "propiedad_id": propiedad.id,
            "modelo": cabeza["MODELO"],
            "alianza_id": catalogos["alianza"].get(cabeza["ALIANZA"] or ""),
            "tipo_operacion_id": catalogos["tipo_operacion"].get(cabeza["TIPO_OPERACION"] or ""),
            "etapa": cabeza["ETAPA"],
            "vendedor_arrendador": cabeza["VENDEDOR_ARRENDADOR"],
            "comprador_arrendatario": cabeza["COMPRADOR_ARRENDATARIO"],
            "corredor_agente": cabeza["CORREDOR_AGENTE"],
            "notas": cabeza["NOTAS"],
        }

        if negocio is None:
            negocio = Negocio(codigo=codigo, **campos_negocio)
            db.add(negocio)
            db.flush()
            resumen.negocios_nuevos += 1
        else:
            for campo, valor in campos_negocio.items():
                setattr(negocio, campo, valor)
            resumen.negocios_actualizados += 1

        existentes = {
            (h.nombre or ""): h
            for h in db.execute(
                select(NegocioHito).where(NegocioHito.negocio_id == negocio.id)
            ).scalars()
        }

        for f in grupo:
            campos_hito = {
                "nombre": f["HITO"],
                "estado": f["ESTADO"],
                "fecha_inicio": f["FECHA_INICIO"],
                "fecha_cierre": f["FECHA_CIERRE"],
                "valor_negocio": f["VALOR_NEGOCIO"],
                "moneda": f["MONEDA"],
                "fecha_valorizacion": f["FECHA_VALORIZACION"],
                "valor_clp_manual": f["VALOR_CLP_MANUAL"],
                "motivo_valor_manual": f["MOTIVO_VALOR_MANUAL"],
                "nombre_tercero": f["NOMBRE_TERCERO"],
                **{c.lower(): f[c] for c in COLUMNAS_PCT},
            }

            hito = existentes.get(f["HITO"] or "")
            if hito is None:
                hito = NegocioHito(negocio_id=negocio.id, **campos_hito)
                db.add(hito)
                resumen.hitos_nuevos += 1
            else:
                for campo, valor in campos_hito.items():
                    setattr(hito, campo, valor)
                resumen.hitos_actualizados += 1

            db.flush()
            # El motor manda: nada de lo que calcula viene del archivo.
            refrescar_hito(db, hito, negocio.modelo)
