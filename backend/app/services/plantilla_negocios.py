"""La plantilla de carga masiva de negocios.

**Una fila es un hito, no un negocio.** Si el código se repite, son varios hitos
del mismo negocio -- como `VVP-3`, que tiene PROMESA y ESCRITURA. Los datos de
nivel negocio (propiedad, modelo, alianza) tienen que coincidir entre esas filas;
si no coinciden es un error, porque no hay forma de saber cuál gana.

**La plantilla pide entradas, no resultados.** No hay columnas para comisión
total, comisión broker ni comisión real VP: eso lo calcula el motor a partir del
valor y las tasas. Ponerlas sería dejar que alguien escriba un número a mano y
perder de un plumazo la garantía que el motor existe para dar.

**Ojo: esto no es la herramienta para los 19 históricos.** Esos se migran fieles
y sin recalcular (`D-026`), porque siete están cerrados con plata ya facturada y
`VVP-2` viene descuadrado del origen. Para eso está
`scripts/cargar_negocios.py`, que migra tal cual y reporta las diferencias contra
el motor. Esta plantilla es para negocios nuevos, donde el motor **debe** mandar.

**Los códigos válidos se escriben en una hoja aparte, leídos de la base.** Nadie
tiene que adivinar si la alianza es `ASSETPLAN` o `Assetplan`, y si mañana se
agrega una alianza la plantilla la trae sola. Es la misma razón por la que los
desplegables del front salen de la API y no de listas escritas en el código.
"""
from dataclasses import dataclass
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalogo import Catalogo, EstadoNegocio, Etapa, ModeloNegocio
from app.services.estructura_archivo import (
    ColumnaArchivo,
    EstructuraArchivo,
    GrupoColumnas,
    ValoresDeColumna,
)

HOJA = "NEGOCIOS"
HOJA_VALORES = "Valores válidos"
HOJA_GUIA = "Instrucciones"


@dataclass(frozen=True)
class Columna:
    nombre: str
    grupo: str
    obligatoria: bool
    ayuda: str
    ancho: int = 18


# El orden es el de trabajo: primero qué negocio es, después la propiedad, el
# estado, la plata y al final las tasas. Quien llena la fila la lee de izquierda
# a derecha en el mismo orden en que tiene los datos a mano.
COLUMNAS: tuple[Columna, ...] = (
    Columna("CODIGO", "Identificación", True, "El código del negocio, por ejemplo VVP-20. Repetirlo agrega otro hito al mismo negocio.", 14),
    Columna("HITO", "Identificación", False, "Nombre del hito cuando el negocio tiene más de uno: PROMESA, ESCRITURA. Vacío si es uno solo.", 14),

    Columna("DIRECCION", "Propiedad", True, "Calle y número.", 34),
    Columna("UNIDAD", "Propiedad", False, "Departamento, oficina o bodega.", 12),
    Columna("COMUNA", "Propiedad", True, "Comuna de la propiedad.", 18),
    Columna("TIPO_PROPIEDAD", "Propiedad", False, "Código del catálogo de tipos de propiedad.", 18),

    Columna("MODELO", "Negocio", True, "Modelo de negocio. Define cómo se reparte la comisión, así que no puede quedar vacío.", 26),
    Columna("ALIANZA", "Negocio", False, "Código de la alianza, si el negocio viene de una.", 16),
    Columna("TIPO_OPERACION", "Negocio", False, "VENTA o ARRIENDO.", 16),
    Columna("ETAPA", "Negocio", False, "Etapa del pipeline, E1 a E7. La etapa es del negocio, no del hito.", 10),
    Columna("VENDEDOR_ARRENDADOR", "Negocio", False, "Quién vende o arrienda.", 26),
    Columna("COMPRADOR_ARRENDATARIO", "Negocio", False, "Quién compra o arrienda.", 26),
    Columna("CORREDOR_AGENTE", "Negocio", False, "Corredor o agente de la otra parte.", 24),

    Columna("ESTADO", "Hito", True, "ACTIVO, CERRADO, PERDIDO o DESISTIDO.", 14),
    Columna("FECHA_INICIO", "Hito", True, "Cuándo empezó. Formato 2026-08-21 o 21-08-2026.", 14),
    Columna("FECHA_CIERRE", "Hito", False, "Cuándo se cerró. Solo para los CERRADO.", 14),

    Columna("VALOR_NEGOCIO", "Valorización", False, "El monto del negocio, en la moneda de la columna siguiente.", 16),
    Columna("MONEDA", "Valorización", False, "UF o CLP.", 10),
    Columna("FECHA_VALORIZACION", "Valorización", False, "Con qué fecha se convierte de UF a pesos. Sin ella no hay conversión posible.", 18),
    Columna("VALOR_CLP_MANUAL", "Valorización", False, "Solo si el valor en pesos no sale de la conversión. Manda sobre el calculado.", 18),
    Columna("MOTIVO_VALOR_MANUAL", "Valorización", False, "Por qué se puso el valor a mano. Opcional, pero conviene dejarlo dicho.", 30),

    Columna("PCT_LADO_VENDEDOR", "Tasas", False, "Comisión del lado vendedor, como porcentaje: 2 es 2%.", 18),
    Columna("PCT_LADO_COMPRADOR", "Tasas", False, "Comisión del lado comprador, como porcentaje.", 18),
    Columna("PCT_BROKER_VENDEDOR", "Tasas", False, "Qué parte se lleva el broker del lado vendedor.", 20),
    Columna("PCT_BROKER_COMPRADOR", "Tasas", False, "Qué parte se lleva el broker del lado comprador.", 20),
    Columna("PCT_VP_VENDEDOR", "Tasas", False, "Qué parte queda para ViveProp del lado vendedor.", 18),
    Columna("PCT_VP_COMPRADOR", "Tasas", False, "Qué parte queda para ViveProp del lado comprador.", 18),
    Columna("PCT_REBATE_CONCENTRADOR", "Tasas", False, "Rebate del concentrador. Solo en el modelo de concentradores.", 22),
    Columna("PCT_EQUIPO", "Tasas", False, "Qué parte de la comisión de ViveProp va al equipo.", 14),
    Columna("PCT_TERCERO", "Tasas", False, "Qué parte va a un tercero, si hay.", 14),
    Columna("NOMBRE_TERCERO", "Tasas", False, "Quién es el tercero.", 22),

    Columna("NOTAS", "Otros", False, "Lo que haga falta anotar.", 34),
)

OBLIGATORIAS = tuple(c.nombre for c in COLUMNAS if c.obligatoria)
NOMBRES = tuple(c.nombre for c in COLUMNAS)

_AZUL = "3D3EA8"     # primario de marca
_CORAL = "F4545A"    # acento: marca las obligatorias
_GRIS = "EDEDF9"


def _encabezar(hoja) -> None:
    """Dos filas de encabezado: el grupo arriba, la columna abajo.

    Con 31 columnas, una sola fila de títulos obliga a contar posiciones para
    saber dónde termina la propiedad y dónde empiezan las tasas.
    """
    grupo_inicio = 1
    for i, col in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=2, column=i, value=col.nombre)
        celda.font = Font(bold=True, color="FFFFFF", size=9)
        celda.fill = PatternFill("solid", fgColor=_CORAL if col.obligatoria else _AZUL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.column_dimensions[get_column_letter(i)].width = col.ancho

        siguiente = COLUMNAS[i].grupo if i < len(COLUMNAS) else None
        if siguiente != col.grupo:
            celda_grupo = hoja.cell(row=1, column=grupo_inicio, value=col.grupo.upper())
            celda_grupo.font = Font(bold=True, size=9, color=_AZUL)
            celda_grupo.alignment = Alignment(horizontal="center")
            if grupo_inicio != i:
                hoja.merge_cells(
                    start_row=1, start_column=grupo_inicio, end_row=1, end_column=i
                )
            grupo_inicio = i + 1

    hoja.row_dimensions[2].height = 42
    # Se congelan las dos filas de encabezado: con 31 columnas se pierde de vista
    # cuál se está llenando en cuanto se baja un poco.
    hoja.freeze_panes = "A3"


def _hoja_valores(libro, db: Session) -> None:
    """Los códigos que la carga acepta, leídos de la base.

    Se genera en vez de escribirse fija para que una alianza nueva aparezca sola
    en la próxima plantilla que alguien baje.
    """
    hoja = libro.create_sheet(HOJA_VALORES)
    hoja.column_dimensions["A"].width = 26
    hoja.column_dimensions["B"].width = 38
    hoja.column_dimensions["C"].width = 12
    fila = 1

    def bloque(titulo: str, filas: list[tuple[str, str]]) -> None:
        nonlocal fila
        celda = hoja.cell(row=fila, column=1, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        hoja.cell(row=fila, column=2).fill = PatternFill("solid", fgColor=_AZUL)
        fila += 1
        if not filas:
            hoja.cell(row=fila, column=1, value="(no hay ninguno cargado)").font = Font(italic=True)
            fila += 2
            return
        for codigo, nombre in filas:
            hoja.cell(row=fila, column=1, value=codigo).font = Font(name="Consolas")
            hoja.cell(row=fila, column=2, value=nombre)
            fila += 1
        fila += 1

    bloque("MODELO", [(m.value, m.value.replace("_", " ").title()) for m in ModeloNegocio])
    bloque("ESTADO", [(e.value, e.value.title()) for e in EstadoNegocio])
    bloque("MONEDA", [("UF", "Unidad de Fomento"), ("CLP", "Pesos chilenos")])

    etapas = db.execute(select(Etapa).order_by(Etapa.orden)).scalars().all()
    bloque("ETAPA", [(e.codigo, e.nombre) for e in etapas])

    for tipo, titulo in (
        ("alianza", "ALIANZA"),
        ("tipo_operacion", "TIPO_OPERACION"),
        ("tipo_propiedad", "TIPO_PROPIEDAD"),
    ):
        filas = db.execute(
            select(Catalogo)
            .where(Catalogo.tipo == tipo, Catalogo.activo.is_(True))
            .order_by(Catalogo.orden)
        ).scalars().all()
        bloque(titulo, [(c.codigo, c.nombre) for c in filas])


def _hoja_guia(libro) -> None:
    hoja = libro.create_sheet(HOJA_GUIA)
    hoja.column_dimensions["A"].width = 30
    hoja.column_dimensions["B"].width = 96

    lineas: list[tuple[str, str]] = [
        ("Cómo se llena", ""),
        ("Una fila es un hito", "Si repites el CODIGO, agregas otro hito al mismo negocio. Los datos del negocio "
                               "(propiedad, modelo, alianza) tienen que ser iguales en esas filas."),
        ("Las coral son obligatorias", "Las columnas con encabezado coral no pueden quedar vacías: "
                                       + ", ".join(OBLIGATORIAS) + "."),
        ("Los códigos van exactos", f"Están todos en la hoja «{HOJA_VALORES}», sacados de la base. Cópialos de ahí."),
        ("Las tasas van en porcentaje", "Escribe 2 para 2%, o 2,52 para 2,52%. No 0,02."),
        ("Las fechas", "2026-08-21 o 21-08-2026. También sirve una celda con formato de fecha de Excel."),
        ("", ""),
        ("Qué NO se llena", ""),
        ("Las comisiones no se escriben", "No hay columnas de comisión total, comisión broker ni comisión real VP: "
                                          "las calcula el sistema con el valor y las tasas. Es la razón de que exista "
                                          "el motor de comisiones."),
        ("El valor en pesos tampoco", "Sale de convertir el valor en UF con la fecha de valorización. "
                                      "VALOR_CLP_MANUAL es solo para el caso excepcional en que ese cálculo no aplica."),
        ("", ""),
        ("Qué pasa al cargar", ""),
        ("Si hay errores no se carga nada", "Se revisan todas las filas y se informan todos los problemas juntos. "
                                            "Media carga es peor que ninguna: nadie sabe cuál mitad quedó."),
        ("Cargar dos veces no duplica", "Un código que ya existe se actualiza. Se puede corregir el archivo y volver a subirlo."),
        ("La propiedad se reutiliza", "Si la dirección y la comuna ya existen, se usa esa propiedad en vez de crear una repetida."),
    ]

    fila = 1
    for izquierda, derecha in lineas:
        if izquierda and not derecha:  # título de sección
            celda = hoja.cell(row=fila, column=1, value=izquierda)
            celda.font = Font(bold=True, size=12, color=_AZUL)
        elif izquierda:
            celda = hoja.cell(row=fila, column=1, value=izquierda)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(vertical="top")
            derecho = hoja.cell(row=fila, column=2, value=derecha)
            derecho.alignment = Alignment(wrap_text=True, vertical="top")
        fila += 1

    for f in range(1, fila):
        if hoja.cell(row=f, column=2).value:
            hoja.row_dimensions[f].height = 30


def generar_plantilla(db: Session) -> bytes:
    """El .xlsx vacío, con la guía y los códigos válidos de esta base."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = HOJA
    _encabezar(hoja)
    _hoja_guia(libro)
    _hoja_valores(libro, db)

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def estructura_plantilla(db: Session) -> EstructuraArchivo:
    """Lo mismo que pinta la plantilla, para poder mostrarlo en pantalla.

    Se arma desde `COLUMNAS`, que es de donde sale el Excel, así que la pantalla
    no puede quedar describiendo columnas que la plantilla ya no trae. Los valores
    válidos se leen de la base por el mismo motivo que en la hoja «Valores
    válidos»: una alianza nueva aparece sola.
    """
    grupos: list[GrupoColumnas] = []
    for col in COLUMNAS:
        if not grupos or grupos[-1].nombre != col.grupo:
            grupos.append(GrupoColumnas(nombre=col.grupo, columnas=[]))
        grupos[-1].columnas.append(
            ColumnaArchivo(nombre=col.nombre, obligatoria=col.obligatoria, ayuda=col.ayuda)
        )

    def codigos(tipo: str) -> list[str]:
        return [
            c.codigo
            for c in db.execute(
                select(Catalogo)
                .where(Catalogo.tipo == tipo, Catalogo.activo.is_(True))
                .order_by(Catalogo.orden)
            ).scalars()
        ]

    etapas = db.execute(select(Etapa).order_by(Etapa.orden)).scalars().all()

    return EstructuraArchivo(
        titulo="Carga masiva de negocios",
        origen=(
            "La plantilla que se baja acá. Trae los códigos válidos de esta base, "
            "así que no hay que adivinar si la alianza se escribe ASSETPLAN o Assetplan."
        ),
        fila=(
            "Una fila es un hito, no un negocio. Si repetís el CODIGO, agregás otro "
            "hito al mismo negocio, y los datos del negocio --propiedad, modelo, "
            "alianza-- tienen que ser iguales en esas filas."
        ),
        grupos=grupos,
        valores=[
            ValoresDeColumna(columna="MODELO", valores=[m.value for m in ModeloNegocio]),
            ValoresDeColumna(columna="ESTADO", valores=[e.value for e in EstadoNegocio]),
            ValoresDeColumna(columna="MONEDA", valores=["UF", "CLP"]),
            ValoresDeColumna(
                columna="ETAPA",
                valores=[e.codigo for e in etapas],
                nota="La etapa es del negocio, no del hito.",
            ),
            ValoresDeColumna(columna="ALIANZA", valores=codigos("alianza")),
            ValoresDeColumna(columna="TIPO_OPERACION", valores=codigos("tipo_operacion")),
            ValoresDeColumna(columna="TIPO_PROPIEDAD", valores=codigos("tipo_propiedad")),
        ],
        notas=[
            "Las tasas van en porcentaje: 2 es 2%, y 2,52 es 2,52%. No 0,02.",
            "Las fechas van como 2026-08-21 o 21-08-2026. También sirve una celda "
            "con formato de fecha de Excel.",
            "Las comisiones no se escriben. No hay columna de comisión total, de "
            "broker ni de real VP: las calcula el sistema con el valor y las tasas "
            "(D-039). El valor en pesos tampoco, salvo VALOR_CLP_MANUAL: sale de "
            "convertir la UF con la fecha de valorización.",
            "Si el archivo trae un solo error, no se carga nada. Es a propósito: "
            "media carga es peor que ninguna, porque no se sabe qué quedó afuera.",
        ],
    )
