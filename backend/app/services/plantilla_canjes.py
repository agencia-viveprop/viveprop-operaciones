"""La estructura del archivo de canjes, y una plantilla vacía con esos encabezados.

**El archivo no se llena a mano: sale de Dataprop.** Es el resultado de una query
contra su base, con alias fijos que ya se validaron contra la fuente. Por eso las
16 columnas son todas obligatorias --si falta una, el export está mal armado-- y
por eso la plantilla no sirve para tipear canjes, sino para **comparar** cuando la
carga falla y no se entiende por qué.

Los nombres viven en `importar_canjes.COLUMNAS_REQUERIDAS`, que es lo que la carga
verifica de verdad. Acá se les agrega la descripción, y `test_estructura_archivo.py`
exige que las dos listas digan lo mismo: una descripción que hable de una columna
que la carga ya no pide es peor que no tener descripción. Va como test y no como
`assert` de módulo a propósito --un `assert` ahí tumbaría la app al arrancar por un
texto de ayuda desactualizado, que es una respuesta desproporcionada.
"""
from dataclasses import dataclass
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.estructura_archivo import (
    ColumnaArchivo,
    EstructuraArchivo,
    GrupoColumnas,
    ValoresDeColumna,
)
from app.services.importar_canjes import (
    COLUMNAS_REQUERIDAS,
    ESTADO_MAP,
    ETAPA_MAP,
    MONEDA_MAP,
    OPERACION_MAP,
)

HOJA = "CANJES"
HOJA_VALORES = "Valores válidos"

_AZUL = "3D3EA8"
_CORAL = "F4545A"


@dataclass(frozen=True)
class Columna:
    nombre: str
    grupo: str
    ayuda: str
    ancho: int = 20


# **El orden es exactamente el de `COLUMNAS_REQUERIDAS`**, que es el del export.
# Hay un test que lo exige: la plantilla se baja para comparar encabezados contra
# un archivo que falla, y compararlos en distinto orden no sirve de nada.
#
# Los grupos son solo para la pantalla y tienen que ser contiguos, así que salen
# del orden del export y no de lo que uno agruparía a mano: LINK_PROPIEDAD viene
# al final, después del valor, y por eso queda en su propio grupo.
COLUMNAS: tuple[Columna, ...] = (
    Columna("ID_CANJE", "Identificación", "El N° de solicitud de Dataprop. Es la clave: reimportar la misma fila la actualiza, no la duplica.", 12),
    Columna("FECHA_SOLICITUD", "Identificación", "Cuándo se pidió el canje.", 18),
    Columna("FECHA_CIERRE", "Identificación", "Cuándo se cerró. Vacía si sigue abierto.", 18),
    Columna("ESTADO", "Identificación", "Activo o Cancelado. Se usa solo al crear: en un canje que ya existe no se toca, porque el estado lo gobierna la app.", 14),
    Columna("ETAPA", "Identificación", "En qué va la solicitud. Igual que el estado, solo se usa al crear.", 22),

    Columna("NOMBRE_CORREDOR_SOLICITANTE", "Corredores", "Quién pide el canje.", 30),
    Columna("NOMBRE_CORREDOR_PROPIETARIO", "Corredores", "Quién tiene la propiedad.", 30),
    Columna("EMAIL_CORREDOR_SOLICITANTE", "Corredores", "Correo del solicitante.", 30),
    Columna("EMAIL_CORREDOR_PROPIETARIO", "Corredores", "Correo del propietario.", 30),

    Columna("TIPO_OPERACION", "Propiedad", "Venta, Arriendo u Otro/Desconocido.", 18),
    Columna("TIPO_PROPIEDAD", "Propiedad", "Texto libre tal como viene de Dataprop: DEPTO, CASA, OFICINA.", 18),
    Columna("COMUNA_PROPIEDAD", "Propiedad", "Comuna.", 20),
    Columna("DIRECCION_PROPIEDAD", "Propiedad", "Calle y número.", 34),

    Columna("VALOR_PROP", "Valor", "El monto publicado, en la moneda de la columna siguiente.", 14),
    Columna("MONEDA_VALOR", "Valor", "CLP, UF u Otra.", 14),

    Columna("LINK_PROPIEDAD", "Enlace", "URL de la ficha en Dataprop.", 34),
)

NOMBRES = tuple(c.nombre for c in COLUMNAS)


def estructura_importacion() -> EstructuraArchivo:
    """Las 16 columnas del export, agrupadas, para mostrarlas en pantalla.

    No recibe sesión: a diferencia de negocios, ningún valor de este archivo sale
    de la base. Los cuatro que no son texto libre son enums fijos, y se leen de
    los mapas que usa la carga en vez de escribirse acá (`D-048`).
    """
    grupos: list[GrupoColumnas] = []
    for col in COLUMNAS:
        if not grupos or grupos[-1].nombre != col.grupo:
            grupos.append(GrupoColumnas(nombre=col.grupo, columnas=[]))
        grupos[-1].columnas.append(
            # Todas obligatorias: si falta una, el export está mal armado y la
            # carga se rechaza entera antes de leer una sola fila.
            ColumnaArchivo(nombre=col.nombre, obligatoria=True, ayuda=col.ayuda)
        )

    return EstructuraArchivo(
        titulo="Importar canjes",
        origen=(
            "El .xlsx que sale de la query contra la base de Dataprop. No se llena "
            "a mano: la plantilla de acá sirve para comparar encabezados cuando la "
            "carga falla y no se entiende por qué."
        ),
        fila="Una fila es una solicitud de canje, identificada por su ID_CANJE.",
        grupos=grupos,
        valores=[
            ValoresDeColumna(
                columna="ESTADO",
                valores=list(ESTADO_MAP),
                nota="Solo se usa al crear el canje. En uno que ya existe no se toca.",
            ),
            ValoresDeColumna(
                columna="ETAPA",
                valores=list(ETAPA_MAP),
                nota="Vacía cae en «Sin etapa». Igual que el estado, solo se usa al crear.",
            ),
            ValoresDeColumna(columna="TIPO_OPERACION", valores=list(OPERACION_MAP)),
            ValoresDeColumna(columna="MONEDA_VALOR", valores=list(MONEDA_MAP)),
        ],
        notas=[
            "Los nombres de las columnas tienen que ser exactos. Si falta alguna, "
            "no se carga nada y el error dice cuál.",
            "Los valores de ESTADO, ETAPA, TIPO_OPERACION y MONEDA_VALOR van tal "
            "como los escribe Dataprop, con acentos y mayúsculas: «En revisión», "
            "no «EN REVISION».",
            "Un canje que ya se está gestionando en la app se ignora completo: ni "
            "sus datos ni su etapa se sobreescriben con el archivo.",
            "Al revés que la carga de negocios, acá una fila mala no frena a las "
            "demás. El export trae cientos de filas de un sistema ajeno, y perder "
            "las 296 buenas por una rara no ayudaría a nadie.",
        ],
    )


def generar_plantilla() -> bytes:
    """El .xlsx vacío con los 16 encabezados exactos y los valores que se aceptan.

    **Los encabezados van en la fila 1, sin fila de grupos**, porque es donde
    `importar_canjes` los busca. La plantilla de negocios sí tiene dos filas --su
    carga lee la segunda-- y copiar ese estilo acá producía una plantilla que su
    propio cargador rechazaba por columnas faltantes. Hay un test que sube esta
    plantilla a la carga y exige que la acepte.

    Además es más fiel: el archivo real es el resultado de una query, y una query
    no devuelve encabezados agrupados. Los grupos existen solo para la pantalla.
    """
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = HOJA

    for i, col in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=i, value=col.nombre)
        celda.font = Font(bold=True, color="FFFFFF", size=9)
        # Todas coral: en este archivo no hay columnas opcionales.
        celda.fill = PatternFill("solid", fgColor=_CORAL)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.column_dimensions[get_column_letter(i)].width = col.ancho

    hoja.row_dimensions[1].height = 42
    hoja.freeze_panes = "A2"

    valores = libro.create_sheet(HOJA_VALORES)
    valores.column_dimensions["A"].width = 24
    valores.column_dimensions["B"].width = 60
    fila = 1
    est = estructura_importacion()

    encabezado = valores.cell(row=fila, column=1, value="COLUMNA")
    encabezado.font = Font(bold=True, color="FFFFFF")
    encabezado.fill = PatternFill("solid", fgColor=_AZUL)
    otro = valores.cell(row=fila, column=2, value="VALORES QUE SE ACEPTAN")
    otro.font = Font(bold=True, color="FFFFFF")
    otro.fill = PatternFill("solid", fgColor=_AZUL)
    fila += 2

    for v in est.valores:
        valores.cell(row=fila, column=1, value=v.columna).font = Font(name="Consolas", bold=True)
        celda = valores.cell(row=fila, column=2, value=" · ".join(v.valores))
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        fila += 1
        if v.nota:
            nota = valores.cell(row=fila, column=2, value=v.nota)
            nota.font = Font(italic=True, size=9)
            nota.alignment = Alignment(wrap_text=True, vertical="top")
            fila += 1
        fila += 1

    for nota in est.notas:
        celda = valores.cell(row=fila, column=2, value=nota)
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        valores.row_dimensions[fila].height = 30
        fila += 1

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
