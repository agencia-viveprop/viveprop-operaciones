"""La plantilla del historial de etapas de negocios, **pre-llenada**.

**Por qué existe.** La vista directorio declara imposible proyectar plazos, y el
motivo es real: no hay un solo movimiento de negocio registrado, y las 7
liquidaciones cerradas del histórico traen la misma fecha de inicio y de cierre
--el Excel de origen traía una sola--. Sin duración ni paso por etapas no se puede
decir *cuándo* va a entrar la plata del pipeline.

Esperar a que se acumule esa historia son meses. Cargarla hacia atrás es una
tarde, y desde que la fecha del avance es editable (`D-066`) el sistema la acepta.
Esta plantilla es para eso.

**Sale pre-llenada, no en blanco.** Trae una fila por cada etapa desde `E1` hasta
la etapa en que cada negocio está hoy, con el código y la etapa ya puestos: solo
hay que escribir fechas. Son 71 filas para los 18 negocios, contra un archivo
vacío donde habría que escribir 213 celdas y no equivocarse en ninguna.

**La secuencia es un supuesto y se dice.** Que un negocio en `E5` haya pasado por
`E1` a `E4` es lo normal, no una certeza. Las filas que no correspondan se borran,
y las que queden sin fecha se ignoran al cargar.

**Dos hojas, porque son dos granos distintos.** El historial es una fila por
etapa; la corrección de fechas de inicio es una fila por **liquidación**, y
`VVP-3` tiene dos. Mezclarlas en una hoja obligaría a repetir el mismo dato en
varias filas y a decidir cuál gana si no coinciden. Separadas, cada fila dice una
cosa y nada más.
"""
from dataclasses import dataclass
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.catalogo import Etapa
from app.models.negocio import Negocio, NegocioHito, clave_de_orden
from app.services.estructura_archivo import (
    ColumnaArchivo,
    EstructuraArchivo,
    GrupoColumnas,
)

HOJA_HISTORIAL = "HISTORIAL"
HOJA_LIQUIDACIONES = "LIQUIDACIONES"
HOJA_GUIA = "Instrucciones"

# La fila del encabezado. Va en 1 y no en 2 como la de negocios: acá no hay
# encabezado superior de grupos que justifique la segunda fila, y una fila de
# menos es una fuente de error de menos --el importador de canjes ya se rechazó a
# sí mismo una vez por leer la hoja en la fila equivocada (`D-046`)--.
FILA_ENCABEZADO = 1


@dataclass(frozen=True)
class Columna:
    nombre: str
    ancho: int
    obligatoria: bool
    ayuda: str
    # Si la escribe el sistema como referencia y el importador la ignora. Van en
    # cursiva y gris en el archivo, para que se lea que no son para llenar.
    referencia: bool = False


COLUMNAS_HISTORIAL: tuple[Columna, ...] = (
    Columna("Negocio", 12, True, "El código, ya puesto. No se cambia."),
    Columna("Etapa", 8, True, "El código de etapa, ya puesto. Borrá la fila si no aplica."),
    Columna(
        "Fecha",
        14,
        False,
        "Cuándo llegó a esa etapa, como dd-mm-aaaa. Vacía = la fila se ignora.",
    ),
    Columna(
        "Descripción",
        46,
        False,
        "Qué pasó en esa etapa. Queda como el comentario del movimiento.",
    ),
    Columna("Inicio registrado", 16, False, "Lo que el sistema tiene hoy. Referencia.", True),
    Columna("Cierre registrado", 16, False, "Lo que el sistema tiene hoy. Referencia.", True),
)

COLUMNAS_LIQUIDACIONES: tuple[Columna, ...] = (
    Columna("Negocio", 12, True, "El código, ya puesto."),
    Columna("Liquidación", 14, True, "Cuál de las liquidaciones del negocio. Ya puesta."),
    Columna(
        "Inicio real",
        14,
        False,
        "La fecha en que empezó de verdad, como dd-mm-aaaa. Vacía = no se toca.",
    ),
    Columna("Inicio registrado", 16, False, "Lo que el sistema tiene hoy. Referencia.", True),
    Columna("Cierre registrado", 16, False, "Lo que el sistema tiene hoy. Referencia.", True),
)

ENCABEZADO = PatternFill("solid", fgColor="3D3EA8")
REFERENCIA = PatternFill("solid", fgColor="EDEDF9")
BLANCO = Font(color="FFFFFF", bold=True)
GRIS = Font(color="6B7280", italic=True)


def _etapas_ordenadas(db: Session) -> list[Etapa]:
    return list(db.scalars(select(Etapa).order_by(Etapa.orden)).all())


def _fecha(valor) -> str:
    return valor.strftime("%d-%m-%Y") if valor else ""


def filas_del_historial(db: Session) -> list[tuple[str, str, str, str, str, str]]:
    """Una fila por etapa alcanzada, con las fechas que el sistema ya tiene.

    Se generan desde `E1` hasta la etapa vigente del negocio. Un negocio sin etapa
    asignada trae solo `E1`: es lo mínimo que se puede afirmar --existe, así que
    alguna vez entró-- y el resto lo agrega quien llena el archivo.
    """
    etapas = _etapas_ordenadas(db)
    orden_de = {e.codigo: e.orden or 0 for e in etapas}

    filas = []
    # Por número: si no, el archivo sale VVP-1, VVP-10, VVP-11 ... VVP-2.
    for negocio in sorted(
        db.scalars(select(Negocio)).all(), key=lambda n: clave_de_orden(n.codigo)
    ):
        # Las fechas de referencia son las del negocio completo: la del hito más
        # antiguo y la del último cierre. Es lo que sirve para orientarse.
        inicios = [h.fecha_inicio for h in negocio.hitos if h.fecha_inicio]
        cierres = [h.fecha_cierre for h in negocio.hitos if h.fecha_cierre]
        hasta = orden_de.get(negocio.etapa or "", 0) or 1

        for etapa in etapas:
            if (etapa.orden or 0) > hasta:
                break
            primera = etapa.orden == min(e.orden or 0 for e in etapas)
            filas.append((
                negocio.codigo,
                etapa.codigo,
                "",
                "",
                # Las de referencia solo en la primera fila del negocio: repetidas
                # en cada fila se leen como si fueran de esa etapa.
                _fecha(min(inicios)) if primera and inicios else "",
                _fecha(max(cierres)) if primera and cierres else "",
            ))
    return filas


def filas_de_liquidaciones(db: Session) -> list[tuple[str, str, str, str, str]]:
    """Una fila por liquidación cuya fecha de inicio es igual a la de cierre.

    Son las que el Excel de origen dejó con una sola fecha, así que su duración es
    desconocida y no se puede calcular nada con ellas. Solo esas: pedir la fecha
    de inicio de las demás sería invitar a cambiar datos que están bien.
    """
    crudas = db.execute(
        select(NegocioHito, Negocio.codigo)
        .join(Negocio, Negocio.id == NegocioHito.negocio_id)
        .where(
            NegocioHito.fecha_cierre.is_not(None),
            NegocioHito.fecha_inicio == NegocioHito.fecha_cierre,
        )
    ).all()

    filas = []
    for hito, codigo in sorted(crudas, key=lambda f: (clave_de_orden(f[1]), f[0].id)):
        filas.append((
            codigo,
            hito.nombre or "ÚNICA",
            "",
            _fecha(hito.fecha_inicio),
            _fecha(hito.fecha_cierre),
        ))
    return filas


def _escribir_hoja(hoja, columnas: tuple[Columna, ...], filas) -> None:
    for i, col in enumerate(columnas, start=1):
        celda = hoja.cell(row=FILA_ENCABEZADO, column=i, value=col.nombre)
        celda.fill = REFERENCIA if col.referencia else ENCABEZADO
        celda.font = GRIS if col.referencia else BLANCO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.column_dimensions[get_column_letter(i)].width = col.ancho

    for f, fila in enumerate(filas, start=FILA_ENCABEZADO + 1):
        for i, valor in enumerate(fila, start=1):
            celda = hoja.cell(row=f, column=i, value=valor)
            if columnas[i - 1].referencia:
                celda.font = GRIS

    hoja.freeze_panes = hoja.cell(row=FILA_ENCABEZADO + 1, column=1)


def _escribir_guia(hoja, columnas_por_hoja) -> None:
    hoja.column_dimensions["A"].width = 22
    hoja.column_dimensions["B"].width = 14
    hoja.column_dimensions["C"].width = 90

    fila = 1
    for titulo, texto in (
        ("Para qué es", "Cargar cuándo cada negocio pasó por cada etapa, y corregir las fechas de inicio que el Excel de origen dejó iguales a la de cierre."),
        ("Qué hay que llenar", "Solo la columna Fecha --y la Descripción si querés-- en la hoja HISTORIAL, y la columna Inicio real en la hoja LIQUIDACIONES. El resto ya viene puesto."),
        ("Las filas sin fecha", "Se ignoran. Si de un negocio solo sabés dos fechas, llená esas dos y dejá el resto vacío."),
        ("Las filas que no aplican", "Borralas. Las etapas vienen pre-generadas suponiendo que un negocio en E5 pasó por E1 a E4; si se salteó alguna, esa fila no corresponde."),
        ("Recargar el archivo", "No duplica. La clave es negocio + etapa: si corregís una fecha y volvés a subir, se actualiza esa fila."),
        ("No agenda nada", "Estos movimientos no generan próxima acción, así que cargar historia no llena «Qué me toca hoy» de vencidos."),
        ("Las columnas grises", "Son lo que el sistema tiene hoy, para que te orientes. No se leen al cargar."),
    ):
        hoja.cell(row=fila, column=1, value=titulo).font = Font(bold=True)
        hoja.cell(row=fila, column=3, value=texto).alignment = Alignment(wrap_text=True)
        fila += 2

    for nombre, columnas in columnas_por_hoja:
        hoja.cell(row=fila, column=1, value=f"Hoja {nombre}").font = Font(bold=True)
        fila += 1
        for col in columnas:
            hoja.cell(row=fila, column=1, value=col.nombre).font = GRIS if col.referencia else None
            hoja.cell(row=fila, column=2, value="referencia" if col.referencia else ("obligatoria" if col.obligatoria else "opcional"))
            hoja.cell(row=fila, column=3, value=col.ayuda).alignment = Alignment(wrap_text=True)
            fila += 1
        fila += 1


def generar_plantilla(db: Session) -> bytes:
    """El .xlsx con las dos hojas pre-llenadas y la guía."""
    libro = openpyxl.Workbook()

    historial = libro.active
    historial.title = HOJA_HISTORIAL
    _escribir_hoja(historial, COLUMNAS_HISTORIAL, filas_del_historial(db))

    liquidaciones = libro.create_sheet(HOJA_LIQUIDACIONES)
    _escribir_hoja(liquidaciones, COLUMNAS_LIQUIDACIONES, filas_de_liquidaciones(db))

    _escribir_guia(
        libro.create_sheet(HOJA_GUIA),
        ((HOJA_HISTORIAL, COLUMNAS_HISTORIAL), (HOJA_LIQUIDACIONES, COLUMNAS_LIQUIDACIONES)),
    )

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def estructura_importacion(db: Session) -> EstructuraArchivo:
    """La misma definición que pinta el Excel, para mostrarla en pantalla.

    Sale de las mismas tuplas, así que la plantilla y la pantalla no pueden decir
    cosas distintas (`D-048`).
    """
    return EstructuraArchivo(
        titulo="Historial de etapas de negocios",
        origen="Una plantilla que se baja pre-llenada desde acá.",
        fila=(
            "En la hoja HISTORIAL, una fila es **una etapa alcanzada por un negocio**: "
            "vienen generadas desde E1 hasta donde está hoy cada negocio, y solo hay que "
            "escribir la fecha. En la hoja LIQUIDACIONES, una fila es una liquidación cuya "
            "fecha de inicio quedó igual a la de cierre en el Excel de origen."
        ),
        grupos=[
            GrupoColumnas(
                nombre=f"Hoja {HOJA_HISTORIAL} — una fila por etapa alcanzada",
                columnas=[
                    ColumnaArchivo(nombre=c.nombre, obligatoria=c.obligatoria, ayuda=c.ayuda)
                    for c in COLUMNAS_HISTORIAL
                ],
            ),
            GrupoColumnas(
                nombre=f"Hoja {HOJA_LIQUIDACIONES} — una fila por liquidación a corregir",
                columnas=[
                    ColumnaArchivo(nombre=c.nombre, obligatoria=c.obligatoria, ayuda=c.ayuda)
                    for c in COLUMNAS_LIQUIDACIONES
                ],
            ),
        ],
        valores=[],
        notas=[
            "Las filas sin fecha se ignoran: si de un negocio solo sabés dos fechas, llená "
            "esas dos y dejá el resto vacío.",
            "Las etapas vienen pre-generadas suponiendo que un negocio en E5 pasó por E1 a "
            "E4. Si se salteó alguna, borrá esa fila.",
            "Recargar el archivo no duplica. La clave es negocio + etapa: corregir una fecha "
            "y volver a subir actualiza esa fila.",
            "Estos movimientos no agendan próxima acción, así que cargar historia no llena "
            "«Qué me toca hoy» de compromisos vencidos.",
            "Las columnas grises son lo que el sistema tiene hoy, para orientarse. No se leen.",
            "La corrección de fechas de inicio no se aplica a una liquidación cuya plata "
            "dependa de esa fecha. Hoy ninguna de las siete depende, y si alguna lo hiciera "
            "se informa en el resumen en vez de moverle el monto.",
        ],
    )
